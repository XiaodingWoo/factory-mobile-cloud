from __future__ import annotations

from io import BytesIO
from pathlib import Path
from html import escape
import re
import os
from textwrap import dedent
from urllib.parse import quote
from uuid import uuid4
import zipfile

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from supabase_config import load_settings
from i18n import language, language_selector, t
from ui_theme import inject_shared_theme
from services.forecast_service import get_latest_forecast_result
from services.monthly_report_service import default_report_month, generate_monthly_report
from config import registered_machine_ids
from generate_qr_codes import (
    OUTPUT_DIR as QR_OUTPUT_DIR,
    build_specs as build_qr_specs,
    create_zip as create_qr_zip,
    make_url as make_qr_url,
    poster_image as qr_poster_image,
    read_env_base_url as read_qr_env_base_url,
    safe_filename_part as safe_qr_filename_part,
    save_poster as save_qr_poster,
)

from data_manager import (
    PERMISSION_COLUMNS,
    archive_completed_production_items,
    authenticate,
    add_loose_goods_record,
    delete_production_record,
    get_change_log,
    get_inventory,
    get_loose_goods,
    get_moulds,
    get_mould_media,
    get_mould_maintenance_audit_log,
    get_mould_maintenance_history,
    get_mould_linked_products,
    get_mould_machine_compatibility,
    get_mould_compatible_machine_ids,
    get_mould_machine_parameter_bundle,
    get_mould_machine_settings_history,
    get_mould_notes_history,
    clean_imported_mould_notes,
    clean_import_note_text,
    get_machine_archive,
    get_product_catalog,
    get_product_mould_links,
    get_primary_mould,
    product_change_alert_payload,
    get_production,
    backfill_production_performance_records,
    calculate_production_performance_metrics,
    get_production_performance_records,
    refresh_production_performance_stock_summary,
    update_production_performance_record,
    get_production_change_requests,
    get_stock_history,
    get_parameter_photo_records,
    update_parameter_photo_review,
    get_user_session,
    get_users,
    add_mould_maintenance_record,
    complete_mould_maintenance_record,
    edit_locked_mould_maintenance_record,
    has_role_permission,
    link_product_to_mould,
    move_production_item,
    apply_production_change_request,
    production_change_apply_preview,
    review_production_change_request,
    stock_in,
    stock_in_from_loose_goods,
    update_loose_goods_status,
    update_production_record,
    upsert_mould_notes,
    set_mould_machine_compatibility,
    save_mould_machine_setting,
    validate_mould_machine_for_production,
    resolve_mould_status,
    role_default_permissions,
    upsert_mould,
    upsert_user,
    unlink_product_from_mould,
)


DEFAULT_STOCK_IN_PIN = os.environ.get("STOCK_IN_PIN", "").strip()


PRODUCTION_ENTRY_STATUSES = ["Running", "Next", "Planned"]
PRODUCTION_STATUS_LABELS = {
    "Running": "Running / 正在生产",
    "Next": "Queued / 队列中",
    "Planned": "Planned / 计划中",
    "Paused": "Paused / 暂停",
    "Finished": "Finished / 完成",
    "Idle": "Idle / 空闲",
    "Completed": "Completed / 完成",
}


def production_status_label(status: object) -> str:
    return localized_status_text(status)


def set_flash(key: str, message: str) -> None:
    st.session_state[key] = message


def show_flash(key: str) -> None:
    message = st.session_state.pop(key, "")
    if message:
        st.success(message)


def inject_css() -> None:
    components.html(
        """
        <script>
        const doc = window.parent.document;
        let viewport = doc.querySelector('meta[name="viewport"]');
        if (!viewport) {
            viewport = doc.createElement('meta');
            viewport.setAttribute('name', 'viewport');
            doc.head.appendChild(viewport);
        }
        viewport.setAttribute('content', 'width=device-width, initial-scale=1.0');

        const tuneMobileInputs = () => {
            doc.querySelectorAll('input[type="password"]').forEach((input) => {
                input.setAttribute('inputmode', 'numeric');
                input.setAttribute('autocomplete', 'one-time-code');
                input.setAttribute('pattern', '[0-9]*');
            });
        };
        tuneMobileInputs();
        new MutationObserver(tuneMobileInputs).observe(doc.body, { childList: true, subtree: true });
        </script>
        """,
        height=0,
        width=0,
    )
    st.markdown(
        """
        <style>
        html, body, .stApp {
            max-width: 100%;
            overflow-x: hidden;
        }
        html, body {
            -webkit-text-size-adjust: 100%;
        }
        * {
            box-sizing: border-box;
        }
        [data-testid="stSidebar"] {
            min-width: 220px;
        }
        [data-testid="stSidebarNav"] {
            display: none !important;
        }
        .stApp {
            background: #f4f6f8;
        }
        .block-container {
            width: 100%;
            padding: 3rem 2rem 4rem;
            max-width: 1440px;
        }
        h1, h2, h3 {
            letter-spacing: 0;
        }
        h1 {
            font-size: 1.9rem !important;
            margin: 0.1rem 0 0.45rem;
            line-height: 1.2;
        }
        h2 {
            font-size: 1.12rem;
            margin-top: 0.8rem;
        }
        h3 {
            font-size: 1rem !important;
        }
        button,
        input,
        textarea,
        [role="button"],
        [role="combobox"] {
            font-size: 16px !important;
        }
        button {
            min-height: 48px !important;
            border-radius: 10px !important;
        }
        div[data-testid="stVerticalBlock"] {
            gap: 0.55rem;
        }
        .stButton > button, .stDownloadButton > button {
            min-height: 52px;
            border-radius: 10px;
            font-size: 1rem;
            font-weight: 650;
            width: 100%;
        }
        .stTextInput input, .stNumberInput input, .stSelectbox div[data-baseweb="select"], .stTextArea textarea,
        .stFileUploader button, .stCameraInput button {
            min-height: 50px;
            font-size: 16px !important;
            border-radius: 10px;
        }
        .stTextInput, .stNumberInput, .stSelectbox, .stTextArea, .stRadio, .stFileUploader, .stCameraInput {
            width: 100%;
            max-width: 100%;
        }
        .stTextArea textarea {
            min-height: 92px;
        }
        .stSelectbox div[data-baseweb="select"] {
            max-width: 100%;
        }
        div[data-testid="stSegmentedControl"] {
            width: max-content !important;
            max-width: 100% !important;
        }
        div[data-testid="stSegmentedControl"] div[role="radiogroup"],
        div[data-testid="stSegmentedControl"] div[role="group"],
        div[data-testid="stSegmentedControl"] div[data-baseweb="button-group"],
        div[data-testid="stSegmentedControl"] > div > div {
            display: grid !important;
            grid-template-columns: repeat(3, 128px) !important;
            gap: 0.75rem !important;
            align-items: stretch !important;
            justify-content: start !important;
            width: max-content !important;
            max-width: 100% !important;
        }
        div[data-testid="stSegmentedControl"] div[role="radiogroup"] > label,
        div[data-testid="stSegmentedControl"] div[role="radiogroup"] > div,
        div[data-testid="stSegmentedControl"] div[role="group"] > label,
        div[data-testid="stSegmentedControl"] div[role="group"] > div,
        div[data-testid="stSegmentedControl"] div[data-baseweb="button-group"] > button,
        div[data-testid="stSegmentedControl"] button {
            width: 128px !important;
            min-width: 128px !important;
            max-width: 128px !important;
            min-height: 48px !important;
            border-radius: 10px !important;
            justify-content: center !important;
            text-align: center !important;
            font-size: 16px !important;
            font-weight: 650 !important;
            margin: 0 !important;
            white-space: normal !important;
            flex: 0 0 128px !important;
        }
        div[data-testid="stSegmentedControl"] label > div,
        div[data-testid="stSegmentedControl"] button > div,
        div[data-testid="stSegmentedControl"] [data-testid="stMarkdownContainer"],
        div[data-testid="stSegmentedControl"] p {
            justify-content: center !important;
            text-align: center !important;
            width: 100% !important;
            white-space: normal !important;
            line-height: 1.15 !important;
            margin: 0 !important;
        }
        @media (max-width: 640px) {
            div[data-testid="stSegmentedControl"] {
                width: 100% !important;
            }
            div[data-testid="stSegmentedControl"] div[role="radiogroup"],
            div[data-testid="stSegmentedControl"] div[role="group"],
            div[data-testid="stSegmentedControl"] div[data-baseweb="button-group"],
            div[data-testid="stSegmentedControl"] > div > div {
                grid-template-columns: repeat(3, minmax(0, 1fr)) !important;
                width: 100% !important;
            }
            div[data-testid="stSegmentedControl"] div[role="radiogroup"] > label,
            div[data-testid="stSegmentedControl"] div[role="radiogroup"] > div,
            div[data-testid="stSegmentedControl"] div[role="group"] > label,
            div[data-testid="stSegmentedControl"] div[role="group"] > div,
            div[data-testid="stSegmentedControl"] div[data-baseweb="button-group"] > button,
            div[data-testid="stSegmentedControl"] button {
                width: 100% !important;
                min-width: 0 !important;
                max-width: none !important;
                flex-basis: auto !important;
            }
        }
        div.st-key-production_catalog_type_i18n div[data-testid="stButtonGroup"] {
            width: max-content !important;
            max-width: 100% !important;
        }
        div.st-key-production_catalog_type_i18n div[data-baseweb="button-group"] {
            display: grid !important;
            grid-template-columns: repeat(3, 128px) !important;
            gap: 0.75rem !important;
            width: max-content !important;
            max-width: 100% !important;
            align-items: stretch !important;
            justify-content: start !important;
            border-collapse: separate !important;
        }
        div.st-key-production_catalog_type_i18n button[data-testid^="stBaseButton-segmented_control"] {
            width: 128px !important;
            min-width: 128px !important;
            max-width: 128px !important;
            min-height: 48px !important;
            border-radius: 10px !important;
            margin: 0 !important;
            justify-content: center !important;
            text-align: center !important;
            white-space: normal !important;
            flex: 0 0 128px !important;
        }
        div.st-key-production_catalog_type_i18n button[data-testid^="stBaseButton-segmented_control"] > div,
        div.st-key-production_catalog_type_i18n button[data-testid^="stBaseButton-segmented_control"] span,
        div.st-key-production_catalog_type_i18n button[data-testid^="stBaseButton-segmented_control"] [data-testid="stMarkdownContainer"],
        div.st-key-production_catalog_type_i18n button[data-testid^="stBaseButton-segmented_control"] p {
            width: 100% !important;
            justify-content: center !important;
            text-align: center !important;
            line-height: 1.15 !important;
            margin: 0 !important;
            white-space: normal !important;
        }
        @media (max-width: 640px) {
            div.st-key-production_catalog_type_i18n div[data-testid="stButtonGroup"] {
                width: 100% !important;
            }
            div.st-key-production_catalog_type_i18n div[data-baseweb="button-group"] {
                grid-template-columns: repeat(3, minmax(0, 1fr)) !important;
                width: 100% !important;
            }
            div.st-key-production_catalog_type_i18n button[data-testid^="stBaseButton-segmented_control"] {
                width: 100% !important;
                min-width: 0 !important;
                max-width: none !important;
                flex-basis: auto !important;
            }
        }
        div[data-testid="stForm"] {
            border: 1px solid #d9dee7;
            border-radius: 10px;
            padding: 0.78rem;
            margin-bottom: 5.5rem;
            background: #ffffff;
        }
        div[data-testid="stFormSubmitButton"] {
            position: relative;
            z-index: 1;
            padding-top: 0.35rem;
            padding-bottom: 0.2rem;
            margin-bottom: 5rem;
            background: #ffffff;
        }
        div[data-testid="stFormSubmitButton"] button {
            min-height: 56px;
            width: 100%;
            border-radius: 12px;
            background: #16a34a;
            color: #ffffff;
            border: 0;
            box-shadow: 0 6px 18px rgba(22, 163, 74, 0.22);
        }
        .metric-grid {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.55rem;
            margin: 0.65rem 0;
        }
        .info-card, .metric-card {
            background: #ffffff;
            border: 1px solid #d9dee7;
            border-radius: 10px;
            padding: 0.78rem;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.05);
        }
        .metric-label {
            color: #6b7280;
            font-size: 0.85rem;
            margin-bottom: 0.25rem;
        }
        .metric-value {
            color: #111827;
            font-size: 1.18rem;
            font-weight: 760;
            overflow-wrap: anywhere;
        }
        .field-label {
            color: #6b7280;
            font-size: 0.78rem;
            margin-top: 0.36rem;
        }
        .field-value {
            color: #111827;
            font-size: 0.98rem;
            font-weight: 650;
            overflow-wrap: anywhere;
        }
        .weight-panel {
            background: #f8fafc;
            border: 1px solid #cbd5e1;
            border-radius: 10px;
            padding: 0.72rem;
            margin-top: 0.65rem;
        }
        .weight-panel .field-label:first-child {
            margin-top: 0;
        }
        .note-card {
            background: #fff7ed;
            border: 1px solid #fed7aa;
            border-radius: 8px;
            padding: 0.72rem;
            margin-top: 0.6rem;
            color: #7c2d12;
            font-size: 0.95rem;
            font-weight: 620;
            overflow-wrap: anywhere;
        }
        .mould-note-popover {
            margin-top: 0.36rem;
        }
        .mould-note-popover summary {
            cursor: pointer;
            list-style: none;
            border: 1px solid #cbd5e1;
            border-radius: 10px;
            background: #f8fafc;
            padding: 0.55rem 0.65rem;
            transition: border-color 0.15s ease, background 0.15s ease;
        }
        .mould-note-popover summary::-webkit-details-marker {
            display: none;
        }
        .mould-note-popover[open] summary {
            border-color: #2563eb;
            background: #eef6ff;
        }
        .mould-note-hint {
            display: inline-block;
            margin-top: 0.3rem;
            color: #2563eb;
            font-size: 0.78rem;
            font-weight: 700;
        }
        .mould-note-bubble {
            margin-top: 0.45rem;
            border: 1px solid #fed7aa;
            border-left: 4px solid #f97316;
            background: #fff7ed;
            color: #7c2d12;
            border-radius: 10px;
            padding: 0.7rem;
            white-space: pre-wrap;
            overflow-wrap: anywhere;
            box-shadow: 0 8px 22px rgba(124, 45, 18, 0.12);
        }
        .product-card {
            margin: 0.65rem 0;
            border-left: 4px solid #2563eb;
        }
        .status-running {
            border-left-color: #16a34a;
        }
        .status-stopped, .status-paused, .status-completed {
            border-left-color: #dc2626;
        }
        .status-setup, .status-changeover, .status-next, .status-planned {
            border-left-color: #f59e0b;
        }
        .status-maintenance {
            border-left-color: #64748b;
        }
        .status-pill {
            flex: 0 0 auto;
            border-radius: 999px;
            padding: 0.28rem 0.58rem;
            font-size: 0.82rem;
            font-weight: 800;
            line-height: 1.15;
            white-space: nowrap;
        }
        .status-pill.status-running {
            background: #dcfce7;
            border: 1px solid #86efac;
            color: #166534;
        }
        .status-pill.status-stopped, .status-pill.status-paused, .status-pill.status-completed {
            background: #fee2e2;
            border: 1px solid #fecaca;
            color: #991b1b;
        }
        .status-pill.status-setup, .status-pill.status-changeover, .status-pill.status-next, .status-pill.status-planned {
            background: #fef3c7;
            border: 1px solid #fde68a;
            color: #92400e;
        }
        .status-pill.status-maintenance {
            background: #e0f2fe;
            border: 1px solid #bae6fd;
            color: #075985;
        }
        .machine-overview-card {
            display: block;
            color: inherit;
        }
        .machine-number {
            color: #111827;
            font-size: 1.2rem;
            font-weight: 850;
            line-height: 1.1;
        }
        .machine-meta-grid {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.5rem;
            margin-top: 0.65rem;
        }
        .machine-meta {
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            border-radius: 9px;
            padding: 0.56rem;
            min-width: 0;
        }
        .machine-meta-label {
            color: #6b7280;
            font-size: 0.82rem;
            margin-bottom: 0.2rem;
        }
        .machine-meta-value {
            color: #111827;
            font-size: 1rem;
            font-weight: 760;
            overflow-wrap: anywhere;
        }
        .product-head {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            gap: 0.7rem;
            margin-bottom: 0.5rem;
        }
        .product-index {
            color: #2563eb;
            font-size: 0.85rem;
            font-weight: 760;
            margin-bottom: 0.18rem;
        }
        .product-title {
            color: #111827;
            font-size: 1.04rem;
            line-height: 1.25;
            font-weight: 780;
            overflow-wrap: anywhere;
        }
        .progress-pill {
            flex: 0 0 auto;
            background: #eef2ff;
            border: 1px solid #c7d2fe;
            color: #3730a3;
            border-radius: 999px;
            padding: 0.24rem 0.5rem;
            font-size: 0.78rem;
            font-weight: 760;
        }
        .compact-metrics {
            grid-template-columns: repeat(3, minmax(0, 1fr));
            gap: 0.5rem;
        }
        .compact-metrics .metric-card {
            padding: 0.54rem;
        }
        .compact-metrics .metric-value {
            font-size: 0.98rem;
        }
        .tag-row {
            display: flex;
            flex-wrap: wrap;
            gap: 0.45rem;
            margin-top: 0.7rem;
        }
        .tag {
            display: inline-block;
            background: #f3f4f6;
            border: 1px solid #e5e7eb;
            border-radius: 999px;
            color: #374151;
            padding: 0.28rem 0.55rem;
            font-size: 0.8rem;
            line-height: 1.25;
            max-width: 100%;
            overflow-wrap: anywhere;
        }
        .production-notes-card {
            margin-top: 0.8rem;
            border: 1px solid #d8e3f3;
            border-radius: 10px;
            overflow: hidden;
            background: #ffffff;
        }
        .production-notes-title {
            padding: 0.55rem 0.72rem;
            font-size: 0.95rem;
            font-weight: 850;
            color: #1e3a8a;
            background: #eff6ff;
            border-bottom: 1px solid #d8e3f3;
        }
        .production-notes-table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
            font-size: 0.9rem;
        }
        .production-notes-table th,
        .production-notes-table td {
            border-bottom: 1px solid #e5edf7;
            padding: 0.48rem 0.55rem;
            vertical-align: top;
            overflow-wrap: anywhere;
        }
        .production-notes-table tr:last-child th,
        .production-notes-table tr:last-child td {
            border-bottom: 0;
        }
        .production-notes-group {
            width: 24%;
            font-weight: 850;
            text-align: center;
        }
        .production-notes-field {
            width: 38%;
            font-weight: 780;
            color: #1f2937;
        }
        .production-notes-value {
            width: 38%;
            color: #111827;
            font-weight: 500;
        }
        .production-alert-note {
            margin: 0.65rem 0.72rem;
            padding: 0.65rem 0.75rem;
            border: 2px solid #dc2626;
            border-radius: 10px;
            background: #fff1f2;
            color: #991b1b;
            font-weight: 850;
            line-height: 1.42;
            overflow-wrap: anywhere;
        }
        .production-alert-note .alert-title {
            font-size: 0.98rem;
            margin-bottom: 0.3rem;
        }
        .production-alert-note .alert-body {
            white-space: pre-line;
            font-size: 0.92rem;
        }
        .notes-packaging {
            background: #dbeafe;
            color: #1d4ed8;
        }
        .notes-spec {
            background: #dcfce7;
            color: #166534;
        }
        .notes-protection {
            background: #ffedd5;
            color: #c2410c;
        }
        .notes-qc {
            background: #fff1b8;
            color: #7c2d12;
        }
        .production-notes-table tr.notes-qc th,
        .production-notes-table tr.notes-qc td,
        .production-notes-table th.notes-qc,
        .production-notes-table td.notes-qc {
            background: #fff1b8 !important;
            color: #7c2d12 !important;
            border-color: #f97316 !important;
            font-weight: 850;
        }
        .production-notes-table tr.notes-qc .production-notes-group,
        .production-notes-table th.production-notes-group.notes-qc,
        .production-notes-group.notes-qc {
            background: #f97316 !important;
            color: #ffffff !important;
            border-color: #ea580c !important;
            font-weight: 900;
        }
        .production-notes-table tr.notes-qc .production-notes-value,
        .production-notes-table td.production-notes-value.notes-qc {
            color: #991b1b !important;
            font-weight: 900;
        }
        .mobile-nav {
            position: fixed;
            left: 0;
            right: 0;
            bottom: 0;
            z-index: 999;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(72px, 1fr));
            gap: 0.25rem;
            padding: 0.42rem 0.5rem calc(0.42rem + env(safe-area-inset-bottom));
            background: rgba(255, 255, 255, 0.96);
            border-top: 1px solid #d9dee7;
            box-shadow: 0 -8px 20px rgba(15, 23, 42, 0.08);
        }
        .mobile-nav:has(a:only-child) {
            grid-template-columns: 1fr;
        }
        .mobile-nav a {
            text-decoration: none;
            color: #4b5563;
            text-align: center;
            font-size: 0.78rem;
            font-weight: 700;
            padding: 0.58rem 0.28rem;
            border-radius: 10px;
            min-width: 0;
        }
        .mobile-nav a.active {
            color: #ffffff;
            background: #2563eb;
        }
        .top-strip {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 0.6rem;
            padding: 0.45rem 0.55rem;
            margin-bottom: 0.45rem;
            background: #ffffff;
            border: 1px solid #d9dee7;
            border-radius: 8px;
            color: #374151;
            font-size: 0.82rem;
            font-weight: 650;
        }
        .quick-summary {
            position: sticky;
            top: 0;
            z-index: 5;
            background: #f4f6f8;
            padding-top: 0.2rem;
        }
        .stock-selected-card {
            border-left: 4px solid #16a34a;
            margin: 0.55rem 0;
        }
        .stock-in-route {
            display: none;
        }
        .stApp:has(.stock-in-route) .mobile-nav {
            position: static;
            margin-top: 1rem;
            margin-bottom: calc(0.4rem + env(safe-area-inset-bottom));
            border: 1px solid #d9dee7;
            border-radius: 10px;
            box-shadow: none;
        }
        .stApp:has(.stock-in-route) .block-container {
            padding-bottom: 1rem;
        }
        .stApp:has(.stock-in-route) div[data-testid="stForm"] {
            margin-bottom: 0.8rem;
        }
        .stApp:has(.stock-in-route) div[data-testid="stFormSubmitButton"] {
            position: sticky;
            bottom: calc(0.45rem + env(safe-area-inset-bottom));
            z-index: 20;
            margin: 0.45rem 0 0.35rem;
            background: linear-gradient(180deg, rgba(255, 255, 255, 0.72), #ffffff 35%);
        }
        .stock-selected-title {
            color: #111827;
            font-size: 1.08rem;
            font-weight: 820;
            line-height: 1.25;
            overflow-wrap: anywhere;
        }
        .stock-photo-note {
            color: #6b7280;
            font-size: 0.9rem;
        }
        .forecast-summary-card {
            background: #ffffff;
            border: 1px solid #d9dee7;
            border-left: 5px solid #16a34a;
            border-radius: 10px;
            padding: 0.82rem;
            margin: 0.4rem 0 0.8rem;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.05);
        }
        .forecast-summary-card.forecast-yellow {
            border-left-color: #f59e0b;
        }
        .forecast-summary-card.forecast-red {
            border-left-color: #dc2626;
        }
        .forecast-badge {
            display: inline-flex;
            min-height: 30px;
            align-items: center;
            border-radius: 999px;
            padding: 0.22rem 0.6rem;
            font-size: 0.82rem;
            font-weight: 850;
            background: #dcfce7;
            color: #166534;
            border: 1px solid #86efac;
        }
        .forecast-badge.forecast-yellow {
            background: #fef3c7;
            color: #92400e;
            border-color: #fde68a;
        }
        .forecast-badge.forecast-red {
            background: #fee2e2;
            color: #991b1b;
            border-color: #fecaca;
        }
        .stCameraInput button, .stFileUploader button {
            min-height: 52px;
            border-radius: 12px;
            font-size: 16px !important;
            font-weight: 760;
        }
        .stFileUploader section {
            border-radius: 12px;
            padding: 0.7rem;
        }
        iframe {
            max-width: 100%;
        }
        div[data-testid="stExpander"] {
            background: #ffffff;
            border-radius: 10px;
        }
        @media (max-width: 768px) {
            .block-container {
                max-width: 100%;
                padding: 2.75rem 1rem 4rem;
            }
            .compact-metrics {
                grid-template-columns: repeat(2, minmax(0, 1fr));
            }
            .product-head {
                gap: 0.5rem;
            }
            .top-strip {
                margin-bottom: 0.35rem;
            }
            div[data-testid="stForm"] {
                padding: 0.7rem;
            }
        }
        @media (max-width: 520px) {
            .metric-grid {
                grid-template-columns: repeat(2, minmax(0, 1fr));
            }
            .compact-metrics {
                grid-template-columns: repeat(2, minmax(0, 1fr));
            }
            .block-container {
                padding-left: 12px;
                padding-right: 12px;
            }
            .stRadio div[role="radiogroup"] {
                gap: 0.35rem;
            }
            .stRadio label {
                min-height: 48px;
                align-items: center;
                font-size: 16px;
            }
            div[data-testid="column"] {
                width: 100% !important;
                flex: 1 1 100% !important;
            }
            .machine-meta-grid {
                grid-template-columns: 1fr;
            }
            .production-notes-table,
            .production-notes-table tbody,
            .production-notes-table tr,
            .production-notes-table th,
            .production-notes-table td {
                display: block;
                width: 100%;
            }
            .production-notes-group {
                text-align: left;
                border-bottom: 1px solid rgba(255, 255, 255, 0.6);
            }
            .production-notes-field {
                padding-bottom: 0.12rem;
            }
            .production-notes-value {
                padding-top: 0.12rem;
            }
            h1,
            div[data-testid="stMarkdownContainer"] h1 {
                font-size: 1.65rem !important;
                line-height: 1.18 !important;
            }
        }
        @media (max-width: 380px) {
            .metric-grid,
            .compact-metrics {
                grid-template-columns: 1fr;
            }
            .product-head {
                display: grid;
                grid-template-columns: 1fr;
            }
            .status-pill,
            .progress-pill {
                justify-self: start;
            }
            h1 {
                font-size: 1.5rem !important;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def query_value(name: str, default: str = "") -> str:
    value = st.query_params.get(name, default)
    if isinstance(value, list):
        return value[0] if value else default
    return value or default


def query_flag(name: str) -> bool:
    value = str(query_value(name, "")).strip().lower()
    return value in {"1", "true", "yes", "y", "public"}


def current_user() -> dict[str, object] | None:
    user = st.session_state.get("user")
    if not user:
        return None
    latest = get_user_session(str(user.get("username", "")))
    if not latest:
        st.session_state.pop("user", None)
        return None
    st.session_state["user"] = latest
    return latest


def is_admin() -> bool:
    user = current_user()
    return bool(user and user.get("role") in {"Administrator", "Admin", "Developer"})


def is_developer() -> bool:
    user = current_user()
    return bool(user and user.get("role") == "Developer")


def has_permission(permission: str) -> bool:
    user = current_user()
    if not user:
        return False
    if user.get("role") == "Developer":
        return True
    if permission in user:
        return bool(user.get(permission))
    return has_role_permission(str(user.get("role", "")), permission)


def can_edit_production() -> bool:
    return has_permission("CanEditProduction")


def can_edit_inventory() -> bool:
    return has_permission("CanEditInventory")


def can_edit_moulds() -> bool:
    return has_permission("CanEditMoulds")


def can_edit_mould_notes() -> bool:
    return can_edit_moulds() or has_permission("mould.notes.edit")


def can_edit_mould_compatibility() -> bool:
    return can_edit_moulds() or has_permission("mould.compatibility.edit")


def can_edit_mould_parameters() -> bool:
    return can_edit_moulds() or has_permission("mould.settings.edit")


def can_view_mould_parameters() -> bool:
    return has_permission("mould.settings.view") or has_permission("mould.view")


def can_force_incompatible_mould() -> bool:
    return is_developer() and has_permission("mould.force_incompatible_machine")


def can_manage_users() -> bool:
    return is_developer() or has_permission("CanManageUsers") or has_permission("user.manage")


def can_stock_in() -> bool:
    return has_permission("stock_in.create") or has_permission("CanStockIn") or can_edit_inventory()


def can_manage_loose_goods() -> bool:
    return has_permission("loose_goods.edit") or has_permission("CanManageLooseGoods") or can_edit_inventory() or can_edit_production()


def can_link_product_mould() -> bool:
    return has_permission("mould.link") or has_permission("CanLinkProductMould")


def can_add_mould_maintenance() -> bool:
    return has_permission("maintenance.create") or has_permission("CanAddMouldMaintenance")


def can_complete_mould_maintenance() -> bool:
    return has_permission("maintenance.complete") or has_permission("CanCompleteMouldMaintenance")


def can_edit_locked_mould_maintenance() -> bool:
    return is_developer() or has_permission("maintenance.edit_locked") or has_permission("CanEditLockedMouldMaintenance")


def login_panel() -> None:
    if current_user():
        user = current_user()
        st.markdown(
            f'<div class="top-strip"><span>Factory MIS</span><span>{user["username"]} · {user["role"]}</span></div>',
            unsafe_allow_html=True,
        )
        with st.sidebar:
            st.write(f"Signed in: **{user['username']}**")
            st.write(f"Role: **{user['role']}**")
            language_selector("signed_in_language")
            if st.button(t("login.sign_out")):
                st.session_state.clear()
                st.rerun()
        return

    language_selector("login_language")
    st.title(t("login.title"))
    with st.form("login_form"):
        username = st.text_input(t("login.username"))
        password = st.text_input(t("login.password"), type="password")
        submitted = st.form_submit_button(t("login.submit"))
    if submitted:
        user = authenticate(username, password)
        if user:
            st.session_state["user"] = user
            st.session_state["login_timestamp"] = pd.Timestamp.now().isoformat()
            st.session_state["active_page"] = "machine"
            st.rerun()
        else:
            st.error(t("login.invalid"))
    st.stop()


def public_top_bar() -> None:
    st.markdown(
        '<div class="top-strip"><span>Factory MIS</span><span>Read-only machine view</span></div>',
        unsafe_allow_html=True,
    )


def public_bottom_nav(active_page: str) -> None:
    items = [
        ("machine", "Machine"),
        ("admin", "Login"),
    ]
    links = []
    for page_key, label in items:
        active = " active" if page_key == active_page else ""
        if page_key == "machine":
            href = "?public=1&page=machine"
        else:
            href = "?page=admin"
        links.append(f'<a class="{active}" href="{href}">{label}</a>')
    st.markdown(f'<div class="mobile-nav">{"".join(links)}</div>', unsafe_allow_html=True)


def stock_in_access_allowed(public_mode: bool) -> bool:
    if not public_mode or current_user():
        return True
    if st.session_state.get("stock_in_pin_ok"):
        return True
    st.markdown('<div class="stock-in-route"></div>', unsafe_allow_html=True)
    st.title("Stock-In Access")
    if not DEFAULT_STOCK_IN_PIN:
        st.error("Stock-in PIN is not configured. Set STOCK_IN_PIN before exposing this page.")
        return False
    st.caption("Enter the stock-in PIN to continue.")
    with st.form("stock_in_pin_form"):
        pin = st.text_input("PIN / 入库密码", type="password", placeholder="Enter PIN")
        submitted = st.form_submit_button("Unlock Stock-In")
    if submitted:
        if pin == DEFAULT_STOCK_IN_PIN:
            st.session_state["stock_in_pin_ok"] = True
            st.rerun()
        else:
            st.error("Incorrect PIN.")
    st.info("This page can update inventory. Ask the supervisor for the stock-in PIN.")
    return False


def inventory_lookup(product_code: str) -> dict[str, str]:
    inventory = get_inventory()
    match = inventory[inventory["ProductCode"] == product_code]
    if match.empty:
        return {"unit": "pcs", "location": "", "current_stock": "0", "exists": "No"}
    row = match.iloc[0]
    return {
        "unit": row.get("Unit", "pcs") or "pcs",
        "location": row.get("Location", ""),
        "current_stock": row.get("CurrentStock", "0"),
        "exists": "Yes",
    }


def running_machine_product_selector() -> dict[str, str] | None:
    production = get_production()
    status_series = production["Status"].astype(str).str.strip().str.casefold()
    running = production[status_series == "running"].copy()
    if running.empty:
        st.warning("No running production schedules found. Please start a production schedule before using this source.")
        return None

    machine_ids = sorted({value for value in running["MachineID"].dropna().tolist() if str(value).strip()})
    if not machine_ids:
        st.warning("Running production schedules exist, but no machine number is assigned.")
        return None

    machine_id = st.selectbox("Machine", machine_ids)
    machine_rows = running[running["MachineID"] == machine_id].copy()
    machine_rows["_SortSequence"] = machine_rows["Sequence"].apply(int_value)
    machine_rows = machine_rows.sort_values(["_SortSequence", "ScheduleID"]).reset_index(drop=True)
    choices = [
        f"{row.ProductName or row.ProductCode} | Mould: {row.MouldNumber or '-'} | {row.ColourMasterbatch or '-'}"
        for row in machine_rows.itertuples()
    ]
    if not choices:
        st.warning(f"No running product found for machine {machine_id}.")
        return None

    choice = st.selectbox("Running product", choices)
    selected = machine_rows.iloc[choices.index(choice)].to_dict()
    lookup = inventory_lookup(selected.get("ProductCode", ""))
    return {
        "machine_id": machine_id,
        "schedule_id": selected.get("ScheduleID", ""),
        "production_status": selected.get("Status", ""),
        "product_code": selected.get("ProductCode", ""),
        "product_name": selected.get("ProductName", ""),
        "mould_number": selected.get("MouldNumber", ""),
        "unit": lookup["unit"],
        "location": lookup["location"],
        "current_stock": lookup["current_stock"],
        "create_if_missing": lookup["exists"] == "No",
    }

def nav(default_page: str) -> str:
    page_labels = {
        "machine": t("navigation.machine"),
        "production_table": t("navigation.production"),
        "stock_in": t("navigation.stock"),
        "loose_goods": t("navigation.loose_goods"),
        "moulds": t("navigation.moulds"),
        "product_mould_links": t("navigation.links"),
        "history": t("navigation.history"),
        "admin": t("navigation.admin"),
    }
    available = list(page_labels)
    if not (has_permission("production.view") or has_permission("CanEditProduction")):
        available = [key for key in available if key != "production_table"]
    if not (is_admin() or has_permission("mould.link") or has_permission("CanLinkProductMould")):
        available = [key for key in available if key != "product_mould_links"]
    if not is_admin():
        available = [key for key in available if key != "admin"]
    requested = default_page if default_page in available else ""
    if requested and requested != st.session_state.get("_last_requested_page"):
        initial = requested
        st.session_state["_last_requested_page"] = requested
    else:
        initial = st.session_state.get("active_page", default_page)
    if initial not in available:
        initial = "machine"
    st.session_state["active_page"] = initial
    with st.sidebar:
        selected_label = st.radio(
            "Menu",
            [page_labels[key] for key in available],
            index=available.index(initial),
            key="native_main_navigation",
        )
    selected = available[[page_labels[key] for key in available].index(selected_label)]
    st.session_state["active_page"] = selected
    return selected


def product_mould_links_page() -> None:
    st.title(t("navigation.links"))
    if not can_link_product_mould():
        st.error("Permission denied.")
        return
    catalog = get_product_catalog()
    moulds = get_moulds()
    links = get_product_mould_links()
    keyword = st.text_input("Search product code or name / 搜索产品编号或名称")
    products = catalog.copy()
    if keyword:
        mask = products["Item"].str.contains(keyword, case=False, na=False) | products["ProductDetail"].str.contains(keyword, case=False, na=False)
        products = products[mask]
    if products.empty:
        st.info("No matching products / 没有匹配产品")
        return
    product_options = {
        f"{row.Item} | {row.ProductDetail}": row.Item for row in products.head(200).itertuples()
    }
    product_label = st.selectbox("Product / 产品", list(product_options))
    product_code = product_options[product_label]
    current = links[links["ProductCode"].eq(product_code)]
    st.dataframe(current, use_container_width=True, hide_index=True)
    valid_moulds = moulds[moulds["Active"].str.lower().isin(["true", "yes", "1"])].copy()
    mould_options = {
        f"{row.MouldNumber} | {row.StorageLocation or '-'} | {row.Status or '-'}": row.MouldNumber
        for row in valid_moulds.itertuples()
    }
    if mould_options:
        selected_mould_label = st.selectbox("Registered mould / 已注册模具", list(mould_options))
        primary = st.checkbox("Primary mould / 主模具")
        if st.button("Add link / 添加挂钩", type="primary"):
            try:
                link_product_to_mould(current_user()["username"], product_code, mould_options[selected_mould_label], primary, source_page="ProductMouldLinks", force_primary=True)
                st.success("Link saved / 挂钩已保存")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))
    active_links = current[current["Active"].str.lower().isin(["true", "yes", "1"])]
    if not active_links.empty:
        remove_id = st.selectbox("Deactivate link / 停用挂钩", active_links["LinkID"].tolist())
        confirm = st.checkbox("Confirm deactivation / 确认停用")
        if st.button("Deactivate selected link", disabled=not confirm):
            try:
                unlink_product_from_mould(current_user()["username"], remove_id)
                st.success("Link deactivated.")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))



def show_product_mould_hint(product_code: str, entered_mould: str = "") -> None:
    product_code = str(product_code or "").strip()
    if not product_code:
        return
    primary = get_primary_mould(product_code)
    if not primary:
        st.info("No forecast data yet" if False else "No default mould link yet / 该产品尚未设置默认模具。")
        return
    mould_number = str(primary.get("MouldNumber", "") or "")
    mould_name = str(primary.get("MouldName", "") or primary.get("ProductName", "") or "-")
    status = resolve_mould_status(mould_number)
    st.info(
        f"Linked mould / 已挂钩模具: {mould_number} - {mould_name}; "
        f"Status / 状态: {status.get('status_label')}; Location / 位置: {status.get('location')}"
    )
    entered = str(entered_mould or "").strip()
    if entered and mould_number and entered.casefold() != mould_number.casefold():
        st.warning(
            f"This product currently defaults to mould {mould_number}. "
            f"If you save {entered}, it will be recorded as an additional compatible mould unless an authorized user updates the default."
        )

def int_value(value: object) -> int:
    try:
        return int(float(str(value or 0)))
    except ValueError:
        return 0


def float_value(value: object) -> float:
    try:
        return float(str(value or "").replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def format_weight_g(value: object) -> str:
    number = float_value(value)
    if not number:
        return "Not set"
    return f"{number:,.0f} g/pc" if number.is_integer() else f"{number:,.1f} g/pc"


def catalog_unit_weight_for_row(row: object) -> float:
    get_value = row.get if hasattr(row, "get") else lambda _key, default="": default
    product_code = str(get_value("ProductCode", "") or get_value("Item", "") or "").strip()
    product_name = str(get_value("ProductName", "") or get_value("ProductDetail", "") or "").strip()
    if not product_code and not product_name:
        return 0.0
    try:
        catalog = get_product_catalog()
    except Exception:
        return 0.0
    if catalog.empty:
        return 0.0

    matched = pd.DataFrame()
    if product_code and "Item" in catalog.columns:
        matched = catalog[catalog["Item"].astype(str).str.strip().eq(product_code)]
    if matched.empty and product_name and "ProductDetail" in catalog.columns:
        matched = catalog[catalog["ProductDetail"].astype(str).str.strip().eq(product_name)]
    if matched.empty:
        return 0.0

    catalog_row = matched.iloc[0]
    unit = float_value(catalog_row.get("UnitWeightG", ""))
    if unit:
        return unit
    return (
        float_value(catalog_row.get("MaterialWeightG", ""))
        + float_value(catalog_row.get("SecondMaterialWeightG", ""))
        + float_value(catalog_row.get("MasterbatchWeightG", ""))
    )


def effective_unit_weight_g(row: object) -> float:
    get_value = row.get if hasattr(row, "get") else lambda _key, default="": default
    unit = float_value(get_value("UnitWeightG", ""))
    if unit:
        return unit
    parts_total = (
        float_value(get_value("MaterialWeightG", ""))
        + float_value(get_value("SecondMaterialWeightG", ""))
        + float_value(get_value("MasterbatchWeightG", ""))
    )
    if parts_total:
        return parts_total
    return catalog_unit_weight_for_row(row)


def format_total_weight_kg(unit_weight_g: float, quantity: object) -> str:
    qty = int_value(quantity)
    if not unit_weight_g or not qty:
        return t("common.not_set")
    kg = unit_weight_g * qty / 1000
    return f"{kg:,.2f} kg"


def localized_display_value(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return "-"
    upper = text.upper()
    if upper in {"YES", "Y", "TRUE"}:
        return t("common.yes")
    if upper in {"NO", "N", "FALSE"}:
        return t("common.no")
    if upper in {"N/A", "NA", "NONE", "-"}:
        return t("common.not_applicable")
    if upper == "NOT SET":
        return t("common.not_set")
    return text


def localized_machine_text(machine_id: object) -> str:
    value = str(machine_id or "").strip()
    return f'{t("common.machine")} {value}' if value else "-"


def localized_status_text(status: object) -> str:
    raw = str(status or "").strip()
    if not raw:
        return "-"
    normalized = {
        "running": "Running",
        "run": "Running",
        "next": "Next",
        "queued": "Queued",
        "queue": "Queued",
        "planned": "Planned",
        "paused": "Paused",
        "finished": "Completed",
        "completed": "Completed",
    }.get(raw.casefold(), raw)
    key = f"status.{normalized}"
    translated = t(key)
    return translated if translated != key else raw


def product_weight_block(row: object, quantity_label: str | None = None, quantity: object | None = None) -> str:
    unit_g = effective_unit_weight_g(row)
    return compact_html(
        f'<div class="weight-panel single-weight-panel">'
        f'{card_field(t("production.unit_total_weight"), format_weight_g(unit_g))}'
        f'</div>'
    )


def card_field(label: str, value: object) -> str:
    safe_label = escape(str(label or ""))
    safe_value = escape(localized_display_value(value))
    return f'<div class="field-label">{safe_label}</div><div class="field-value">{safe_value}</div>'


def is_meaningful(value: object, show_no: bool = False) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if not show_no and text.upper() in {"NO", "N", "FALSE", "0", "NONE", "N/A", "-"}:
        return False
    return True


def compact_field(label: str, value: object, show_no: bool = False) -> str:
    if not is_meaningful(value, show_no=show_no):
        return ""
    return card_field(label, value)


def compact_note(label: str, value: object, show_no: bool = False) -> str:
    if not is_meaningful(value, show_no=show_no):
        return ""
    safe_label = escape(str(label or ""))
    safe_value = escape(localized_display_value(value))
    return f'<span class="tag"><b>{safe_label}:</b> {safe_value}</span>'


def additional_packaging_display(value: object) -> str:
    return str(value or "").strip() if is_meaningful(value) else t("common.not_applicable")


def truthy_text(value: object) -> bool:
    return str(value or "").strip().upper() in {"YES", "Y", "TRUE", "1", "是", "需要"}


def qc_required_for_row(row: object) -> bool:
    get_value = row.get if hasattr(row, "get") else lambda _key, default="": default
    return truthy_text(get_value("QCRequired", "")) or is_meaningful(get_value("QCNotes", ""), show_no=True)


def qc_required_display(row: object) -> str:
    return t("common.yes") if qc_required_for_row(row) else t("common.no")


def qc_notes_display(row: object) -> str:
    get_value = row.get if hasattr(row, "get") else lambda _key, default="": default
    notes = str(get_value("QCNotes", "") or "").strip()
    return notes if notes else t("common.not_applicable")


def qc_required_for_save(notes: object) -> bool:
    return is_meaningful(notes, show_no=True)


def qc_notes_for_save(notes: object) -> str:
    return str(notes or "").strip()


def production_change_alert_note(row: pd.Series) -> str:
    status = str(row.get("Status") or "").strip().casefold()
    if status != "running":
        return ""
    payload = product_change_alert_payload(
        row.get("ProductCode", ""),
        row.get("ProductName", ""),
        row.get("MachineID", ""),
        row.get("Status", ""),
    )
    message = _bilingual_alert_text(payload.get("alert_message_en"), payload.get("alert_message_zh"))
    if not message:
        return ""
    title = _bilingual_alert_text(payload.get("alert_title_en"), payload.get("alert_title_zh")) or t("production.change_reminder")
    title = title.replace("\n\n", " / ").replace("\n", " / ")
    machine_id = str(row.get("MachineID") or "-").strip() or "-"
    product_code = str(row.get("ProductCode") or row.get("ProductName") or "-").strip() or "-"
    body = f"{localized_machine_text(machine_id)}  {product_code}\n{message}"
    return (
        '<div class="production-alert-note">'
        f'<div class="alert-title">{escape(title)}</div>'
        f'<div class="alert-body">{escape(body)}</div>'
        '</div>'
    )


def production_notes_table(row: pd.Series) -> str:
    sections = [
        (
            t("production.notes_group_packaging"),
            "notes-packaging",
            [
                (t("production.notes_packaging_type"), row.get("PackagingType")),
                (t("production.notes_packaging_option"), packaging_option_text(row)),
                (t("production.notes_additional_packaging"), additional_packaging_display(row.get("AdditionalPackaging"))),
            ],
        ),
        (
            t("production.notes_group_pack_spec"),
            "notes-spec",
            [
                (t("production.notes_carton_unit_stack"), row.get("CartonUnitStackQty")),
                (t("production.notes_pallet_qty"), row.get("PalletQty")),
                (t("production.notes_pallet_bag"), row.get("PalletBag")),
                (t("production.notes_pallet_type"), row.get("PalletType")),
            ],
        ),
        (
            t("production.notes_group_protection"),
            "notes-protection",
            [
                (t("production.notes_wrap_pallet"), row.get("WrapPallet")),
                (t("production.notes_corner_protector"), row.get("CornerProtector")),
                (t("production.notes_instructions"), row.get("AdditionalInstructions")),
            ],
        ),
        (
            t("production.notes_group_qc"),
            "notes-qc",
            [
                (t("production.qc_details"), qc_notes_display(row)),
            ],
        ),

    ]
    alert_note = production_change_alert_note(row)
    rows: list[str] = []
    for group_label, group_class, pairs in sections:
        visible_pairs = [(label, value) for label, value in pairs if is_meaningful(value, show_no=True)]
        if not visible_pairs:
            continue
        rowspan = len(visible_pairs)
        for index, (field_label, value) in enumerate(visible_pairs):
            group_cell = (
                f'<th class="production-notes-group {group_class}" rowspan="{rowspan}">{escape(group_label)}</th>'
                if index == 0
                else ""
            )
            rows.append(
                f'<tr class="{group_class}">'
                f"{group_cell}"
                f'<td class="production-notes-field {group_class}"><strong>{escape(field_label)}</strong></td>'
                f'<td class="production-notes-value {group_class}">{escape(localized_display_value(value))}</td>'
                "</tr>"
            )
    if not rows and not alert_note:
        return ""
    table_block = ""
    if rows:
        table_block = (
            '<table class="production-notes-table"><tbody>'
            + "".join(rows)
            + '</tbody></table>'
        )
    return (
        '<div class="production-notes-card">'
        + f'<div class="production-notes-title">{escape(t("production.notes_title"))}</div>'
        + alert_note
        + table_block
        + '</div>'
    )


def localized_production_note_text(row: pd.Series, value: object) -> str:
    raw = str(value or "").strip()
    if not is_meaningful(raw, show_no=True):
        return ""
    generated_markers = [
        "Packaging Type:",
        "Packaging Option:",
        "Additional Packaging:",
        "Carton/Unit/Stack:",
        "Pallet Qty:",
        "Pallet Type:",
        "Wrap Pallet:",
        "Food Application:",
        "Corner Protector:",
        "QC Details:",
    ]
    if any(marker in raw for marker in generated_markers):
        pairs = [
            (t("production.notes_packaging_type"), row.get("PackagingType")),
            (t("production.notes_packaging_option"), packaging_option_text(row)),
            (t("production.notes_additional_packaging"), additional_packaging_display(row.get("AdditionalPackaging"))),
            (t("production.notes_carton_unit_stack"), row.get("CartonUnitStackQty")),
            (t("production.notes_pallet_qty"), row.get("PalletQty")),
            (t("production.notes_pallet_type"), row.get("PalletType")),
            (t("production.notes_wrap_pallet"), row.get("WrapPallet")),
            (t("production.food_application"), row.get("FoodApplication")),
            (t("production.notes_corner_protector"), row.get("CornerProtector")),
            (t("production.qc_details"), qc_notes_display(row)),
        ]
        return " | ".join(
            f"{label}: {localized_display_value(item_value)}"
            for label, item_value in pairs
            if is_meaningful(item_value, show_no=True)
        )
    return raw
def compact_html(markup: str) -> str:
    return "".join(line.strip() for line in markup.splitlines() if line.strip())


def status_class(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"running", "run"}:
        return "status-running"
    if text in {"stopped", "stop", "paused", "completed", "finished"}:
        return "status-stopped"
    if text in {"setup", "changeover", "next", "planned"}:
        return "status-setup"
    if text in {"maintenance", "maint", "idle"}:
        return "status-maintenance"
    return "status-maintenance"


def operator_value(row: pd.Series) -> str:
    for column in ["Operator", "CurrentOperator", "AssignedOperator", "User"]:
        value = row.get(column, "")
        if is_meaningful(value, show_no=True):
            return str(value)
    return "-"


def machine_quantity_text(row: pd.Series) -> str:
    planned = int_value(row.get("PlannedQty"))
    completed = int_value(row.get("CompletedQty"))
    if planned:
        return f"{completed:,} / {planned:,}"
    if completed:
        return f"{completed:,}"
    return "0"



def _bilingual_alert_text(en: object, zh: object) -> str:
    en_text = str(en or "").strip()
    zh_text = str(zh or "").strip()
    if language().startswith("zh"):
        return zh_text or en_text
    return en_text or zh_text


def show_running_product_change_alert_once(row: pd.Series, machine_id: object) -> None:
    # Alerts now render inline inside Production Notes to avoid interrupting shop-floor users.
    return

def machine_overview_card(row: pd.Series, machine_id: str) -> str:
    status = row.get("Status") or "No Plan"
    status_css = status_class(status)
    product_title = product_display_title(row) or "No running production plan"
    safe_machine_id = escape(str(machine_id or "-"))
    safe_status = escape(str(status or "No Plan"))
    safe_product_title = escape(str(product_title or "No running production plan"))
    planned = int_value(row.get("PlannedQty"))
    completed = int_value(row.get("CompletedQty"))
    remaining = max(planned - completed, 0)
    progress = completed / planned * 100 if planned else 0
    href = f"?page=machine&machine_id={quote(str(machine_id or ''))}"
    return compact_html(dedent(f"""
        <a href="{href}" class="machine-overview-card" style="text-decoration:none;">
        <div class="info-card product-card {status_css}">
            <div class="product-head">
                <div>
                    <div class="machine-number">{safe_machine_id}</div>
                    <div class="product-title">{safe_product_title}</div>
                </div>
                <div>
                    <div class="status-pill {status_css}">{safe_status}</div>
                    <div class="progress-pill">{progress:.1f}%</div>
                </div>
            </div>
            <div class="metric-grid compact-metrics">
                <div class="metric-card"><div class="metric-label">{escape(t("production.planned"))}</div><div class="metric-value">{planned:,}</div></div>
                <div class="metric-card"><div class="metric-label">{escape(t("production.completed"))}</div><div class="metric-value">{completed:,}</div></div>
                <div class="metric-card"><div class="metric-label">{escape(t("production.remaining"))}</div><div class="metric-value">{remaining:,}</div></div>
            </div>
            <div class="machine-meta-grid">
                <div class="machine-meta">
                    <div class="machine-meta-label">Quantity</div>
                    <div class="machine-meta-value">{machine_quantity_text(row)}</div>
                </div>
                <div class="machine-meta">
                    <div class="machine-meta-label">Last update</div>
                    <div class="machine-meta-value">{row.get("LastUpdated") or "-"}</div>
                </div>
                <div class="machine-meta">
                    <div class="machine-meta-label">Operator</div>
                    <div class="machine-meta-value">{operator_value(row)}</div>
                </div>
                <div class="machine-meta">
                    <div class="machine-meta-label">Mould</div>
                    <div class="machine-meta-value">{row.get("MouldNumber") or "-"}</div>
                </div>
            </div>
            {compact_field(t("production.material"), row.get("Material"))}
            {compact_field(t("production.colour_masterbatch"), row.get("ColourMasterbatch"))}
            {product_weight_block(row)}
            {compact_field("Notes", row.get("Notes"), show_no=True)}
        </div>
        </a>
    """))



def mould_note_popover(mould_number: object, machine_id: object = "") -> str:
    mould_number_text = str(mould_number or "").strip()
    machine_id_text = str(machine_id or "").strip()
    if not mould_number_text:
        return compact_field(t("moulds.number"), mould_number)
    try:
        moulds = get_moulds()
        matches = moulds[moulds["MouldNumber"].astype(str).str.casefold().eq(mould_number_text.casefold())]
    except Exception:
        matches = pd.DataFrame()
    if matches.empty:
        return compact_field(t("moulds.number"), mould_number_text)
    mould = matches.iloc[0]
    sections: list[tuple[str, str]] = []
    mould_note = clean_import_note_text(mould.get("Notes", "")).strip()
    parameter_note = ""
    if machine_id_text:
        try:
            bundle = get_mould_machine_parameter_bundle(mould_number_text, machine_id_text)
            setting = bundle.get("setting") or {}
            if isinstance(setting, dict):
                parameter_note = clean_import_note_text(setting.get("Notes", "")).strip()
        except Exception:
            parameter_note = ""
    maintenance_note = clean_import_note_text(mould.get("MaintenanceNotes", "")).strip()
    if mould_note:
        sections.append((t("moulds.notes_label"), mould_note))
    if parameter_note:
        sections.append((t("moulds.parameter_notes_for_machine", machine=localized_machine_text(machine_id_text)), parameter_note))
    if maintenance_note:
        sections.append((t("moulds.maintenance_notes_label"), maintenance_note))
    if not sections:
        return compact_field(t("moulds.number"), mould_number_text)
    safe_mould = escape(mould_number_text)
    section_html = "".join(
        f'<div class="mould-note-section"><div class="mould-note-section-title">{escape(title)}</div>'
        f'<div class="mould-note-section-body">{escape(body)}</div></div>'
        for title, body in sections
    )
    return compact_html(f"""
        <details class="mould-note-popover">
            <summary>
                <div class="field-label">{escape(t("moulds.number"))}</div>
                <div class="field-value">{safe_mould}</div>
                <span class="mould-note-hint">{escape(t("moulds.click_show_note"))}</span>
            </summary>
            <div class="mould-note-bubble">{section_html}</div>
        </details>
    """)

def product_schedule_card(row: pd.Series, position: int) -> str:
    planned = int_value(row.get("PlannedQty"))
    completed = int_value(row.get("CompletedQty"))
    remaining = max(planned - completed, 0)
    progress = completed / planned * 100 if planned else 0
    title = product_display_title(row)
    status = row.get("Status") or ("Running" if position == 1 else "Planned")
    status_css = status_class(status)
    safe_title = escape(str(title or "Unnamed product"))
    safe_status = escape(localized_status_text(status or "Planned"))
    notes = "".join(
        [
            compact_note(t("production.label_short"), row.get("Label"), show_no=True),
            compact_note(t("production.food_application"), row.get("FoodApplication"), show_no=True),
            compact_note(t("common.location"), row.get("InventoryLocationID")),
        ]
    )
    free_notes = row.get("Notes")
    localized_free_notes = localized_production_note_text(row, free_notes)
    note_block = f'<div class="note-card">{escape(localized_free_notes)}</div>' if localized_free_notes else ""
    tag_block = f'<div class="tag-row">{notes}</div>' if notes else ""
    production_notes = production_notes_table(row)
    return compact_html(dedent(f"""
        <div class="info-card product-card {status_css}">
            <div class="product-head">
                <div>
                    <div class="product-index">#{position} {safe_status}</div>
                    <div class="product-title">{safe_title}</div>
                </div>
                <div class="status-pill {status_css}">{safe_status}</div>
            </div>
            <div class="metric-grid compact-metrics">
                <div class="metric-card"><div class="metric-label">{escape(t("production.planned"))}</div><div class="metric-value">{planned:,}</div></div>
                <div class="metric-card"><div class="metric-label">{escape(t("production.completed"))}</div><div class="metric-value">{completed:,}</div></div>
                <div class="metric-card"><div class="metric-label">{escape(t("production.remaining"))}</div><div class="metric-value">{remaining:,}</div></div>
            </div>
            {compact_field(t("production.last_update"), row.get("LastUpdated"), show_no=True)}
            {compact_field(t("common.operator"), operator_value(row), show_no=True)}
            {compact_field(t("production.product_code_item"), row.get("ProductCode"))}
            {mould_note_popover(row.get("MouldNumber"), row.get("MachineID"))}
            {compact_field(t("production.material"), row.get("Material"))}
            {compact_field(t("production.colour_masterbatch"), row.get("ColourMasterbatch"))}
            {product_weight_block(row)}
            {production_notes}
            {tag_block}
            {note_block}
        </div>
    """))


def archive_completed_for_machine_view(show_message: bool = False) -> int:
    user = current_user()["username"] if current_user() else "system"
    archived_count = archive_completed_production_items(str(user))
    if show_message and archived_count:
        st.success(f"Archived {archived_count} completed production item(s) to MachineArchive.xlsx.")
    return archived_count


def public_machine_overview() -> None:
    archive_completed_for_machine_view(show_message=False)
    production = get_production()
    machine_ids = registered_machine_ids(production)
    if not machine_ids:
        st.title("Machine Status")
        st.info("No machine records found.")
        return

    st.title("Machine Status")
    for machine_id in machine_ids:
        machine_rows = production[production["MachineID"] == machine_id].copy()
        if machine_rows.empty:
            st.markdown(
                f"""
                <div class="info-card">
                    <div class="machine-card-title">Machine {escape(machine_id)}</div>
                    <div class="status-badge">No production plan</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            continue
        running_rows = machine_rows[machine_rows["Status"].str.lower() == "running"].copy()
        display_rows = running_rows if not running_rows.empty else machine_rows.head(1).copy()
        display_rows["_SortSequence"] = display_rows["Sequence"].apply(int_value)
        display_rows = display_rows.sort_values(["_SortSequence", "ScheduleID"])

        for _, row in display_rows.iterrows():
            st.markdown(machine_overview_card(row, machine_id), unsafe_allow_html=True)


def catalog_values_from_schedule(row: pd.Series) -> dict[str, str]:
    return {
        "MachineName": row.get("MachineName", ""),
        "MouldNumber": row.get("MouldNumber", ""),
        "ProductType": row.get("ProductType", ""),
        "Size": row.get("Size", ""),
        "Material": row.get("Material", ""),
        "MaterialLocation": row.get("MaterialLocation", ""),
        "MaterialWeightG": row.get("MaterialWeightG", ""),
        "SecondMaterialWeightG": row.get("SecondMaterialWeightG", ""),
        "MasterbatchWeightG": row.get("MasterbatchWeightG", ""),
        "UnitWeightG": row.get("UnitWeightG", ""),
    }


def normalize_product_type_for_catalog(value: object) -> str:
    text = str(value or "").strip()
    mapping = {"Lid": "盖子", "LID": "盖子", "Bucket": "桶", "BUCKET": "桶", "Pail": "桶", "PAIL": "桶"}
    return mapping.get(text, text)


def packaging_type_from_catalog(row: dict[str, object]) -> str:
    return str(
        row.get("PackagingType")
        or row.get("Packaging Unit")
        or row.get("PackagingUnit")
        or ""
    ).strip()




def value_from_row(row: object, name: str, default: object = "") -> object:
    if hasattr(row, "get"):
        return row.get(name, default)
    return getattr(row, name, default)


def packaging_option_text(row: object) -> str:
    option = str(value_from_row(row, "PackagingOption", "") or "").strip()
    if option:
        return option
    packaging_type = str(value_from_row(row, "PackagingType", "") or "").strip()
    additional = str(value_from_row(row, "AdditionalPackaging", "") or "").strip()
    carton_stack = str(value_from_row(row, "CartonUnitStackQty", "") or "").strip()
    pallet_qty = str(value_from_row(row, "PalletQty", "") or "").strip()
    bits = []
    if packaging_type:
        bits.append(packaging_type)
    if additional and additional.upper() not in {"NO", "N/A", "NONE"}:
        bits.append(additional)
    if carton_stack:
        bits.append(f"Stack/Carton {carton_stack}")
    if pallet_qty:
        bits.append(f"Pallet {pallet_qty}")
    return " - ".join(bits)


def product_display_title(row: object) -> str:
    base = str(
        value_from_row(row, "ProductName", "")
        or value_from_row(row, "ProductDetail", "")
        or value_from_row(row, "ProductCode", "")
        or value_from_row(row, "Item", "")
        or "Unnamed product"
    ).strip()
    option = packaging_option_text(row)
    if option and option.casefold() not in base.casefold():
        return f"{base} - {option}"
    return base

def machine_has_other_running(production: pd.DataFrame, machine_id: object, schedule_id: object = "") -> bool:
    if production.empty:
        return False
    machine_text = str(machine_id or "").strip().casefold()
    schedule_text = str(schedule_id or "").strip()
    mask = (
        production["MachineID"].astype(str).str.strip().str.casefold().eq(machine_text)
        & production["Status"].astype(str).str.strip().str.casefold().eq("running")
    )
    if schedule_text:
        mask = mask & ~production["ScheduleID"].astype(str).eq(schedule_text)
    return bool(mask.any())


def status_options_for_machine(
    production: pd.DataFrame,
    machine_id: object,
    base_options: list[str] | None = None,
    schedule_id: object = "",
) -> list[str]:
    options = list(base_options or PRODUCTION_ENTRY_STATUSES)
    if machine_has_other_running(production, machine_id, schedule_id):
        options = [option for option in options if option.casefold() != "running"]
    return options or ["Next", "Planned"]


def show_running_warning_if_needed(production: pd.DataFrame, machine_id: object, schedule_id: object = "") -> None:
    if machine_has_other_running(production, machine_id, schedule_id):
        st.warning("This machine already has a running product. New product can only be added as Queue or Planning.")


def show_product_detail_native(selected: dict[str, object], notes: str, material_text: str, additive_text: str) -> None:
    st.subheader(t("production.product_details"))
    detail_rows = [
        (t("production.product_detail"), selected.get("ProductDetail", "")),
        (t("production.item"), selected.get("Item", "")),
        (t("production.type"), product_type_filter_label(selected.get("ProductType", ""))),
        (t("production.size"), selected.get("Size", "")),
        (t("production.colour"), selected.get("Colour", "")),
        (t("production.label"), selected.get("HasLabel", "")),
        (t("production.material"), material_text),
        (t("production.additive_masterbatch"), additive_text),
        (t("production.cycle_time"), selected.get("CycleTime", "")),
        (t("production.shot_weight"), selected.get("ShotWeight", "")),
        (t("production.cavity"), selected.get("Cavity", "")),
    ]
    for start in range(0, len(detail_rows), 2):
        columns = st.columns(2)
        for column, (label, value) in zip(columns, detail_rows[start : start + 2]):
            with column:
                st.markdown(f"**{label}**")
                st.write(str(value or "-"))

    st.subheader(t("production.packaging_pallet"))
    packaging_rows = [
        (t("production.packaging_type"), packaging_type_from_catalog(selected)),
        (t("production.packaging_option"), packaging_option_text(selected)),
        (t("production.additional_packaging"), additional_packaging_display(selected.get("AdditionalPackaging"))),
        (t("production.carton_unit_stack"), selected.get("CartonUnitStackQty", "")),
        (t("production.pallet_type"), selected.get("PalletType", "")),
        (t("production.pallet_qty"), selected.get("PalletQty", "")),
        (t("production.packaging_notes"), notes or t("production.no_additional_instructions")),
    ]
    for label, value in packaging_rows:
        if is_meaningful(value, show_no=True):
            st.markdown(f"**{label}:** {value}")


def add_temp_plan_panel(machine_id: str, machine_rows: pd.DataFrame) -> None:
    if not can_edit_production():
        return
    show_flash(f"temp_plan_flash_{machine_id}")
    current_row = machine_rows[machine_rows["Status"].str.lower() == "running"]
    base_row = current_row.iloc[0] if not current_row.empty else machine_rows.iloc[0]
    base = catalog_values_from_schedule(base_row)
    catalog = get_product_catalog()
    if catalog.empty:
        return

    st.subheader("Temporary Add Plan / 临时增产")
    st.caption("Current version uses the current mould number, product type, and size as a soft same-mould filter. Later this can be linked directly to the mould register.")
    st.markdown(
        f"""
        <div class="info-card">
            {card_field("Current mould", base.get("MouldNumber"))}
            {compact_field("Current type", base.get("ProductType"))}
            {compact_field("Current size", base.get("Size"))}
        </div>
        """,
        unsafe_allow_html=True,
    )

    filtered = catalog.copy()
    product_type = normalize_product_type_for_catalog(base.get("ProductType"))
    if is_meaningful(product_type):
        filtered = filtered[filtered["ProductType"] == product_type]
    if is_meaningful(base.get("Size")):
        filtered = filtered[filtered["Size"] == base["Size"]]

    keyword = st.text_input("Search compatible product / 搜索可增产产品", placeholder="Colour, detail, material", key=f"temp_keyword_{machine_id}")
    if keyword:
        words = [word.strip() for word in keyword.replace(",", " ").split() if word.strip()]
        for word in words:
            mask = pd.Series(False, index=filtered.index)
            for column in ["ProductDetail", "Item", "Colour", "MainMaterial", "Additive"]:
                mask = mask | filtered[column].astype(str).str.contains(word, case=False, na=False, regex=False)
            filtered = filtered[mask]

    if filtered.empty:
        st.warning("No compatible products found under the current soft mould filter.")
        return

    filtered = filtered.reset_index(drop=True)
    choices = [product_choice_label(row, i) for i, row in enumerate(filtered.itertuples())]
    status_options = status_options_for_machine(machine_rows, machine_id, PRODUCTION_ENTRY_STATUSES)
    show_running_warning_if_needed(machine_rows, machine_id)
    with st.form(f"temp_plan_form_{machine_id}"):
        selected = st.selectbox("Product / 产品", choices)
        sequence_default = max([int_value(value) for value in machine_rows["Sequence"].tolist()] + [0]) + 1
        sequence = st.number_input("Sequence / 顺序", min_value=1, step=1, value=sequence_default)
        planned_qty = st.number_input("Temporary planned quantity / 临时计划数量", min_value=1, step=1)
        qc_notes = st.text_input(
            t("production.qc_details"),
            placeholder=t("production.qc_details_placeholder"),
            key=f"temp_qc_notes_{machine_id}",
        )
        status = st.selectbox(
            "Status / 状态",
            status_options,
            index=0,
            format_func=production_status_label,
        )
        extra_notes = st.text_area("Reason / Notes / 原因或备注")
        submitted = st.form_submit_button("Add Temporary Plan / 添加临时计划")

    if submitted:
        selected_catalog = filtered.iloc[choices.index(selected)].to_dict()
        notes = "Temporary add plan"
        if extra_notes:
            notes = f"{notes} | {extra_notes}"
        try:
            schedule_id = update_production_record(
                current_user()["username"],
                machine_id,
                {
                    "MachineName": base.get("MachineName", ""),
                    "Sequence": str(sequence),
                    "Status": status,
                    "ProductType": selected_catalog.get("ProductType", ""),
                    "Size": selected_catalog.get("Size", ""),
                    "ProductCode": selected_catalog.get("Item", ""),
                    "ProductName": selected_catalog.get("ProductDetail", "") or selected_catalog.get("Item", ""),
                    "PlannedQty": str(planned_qty),
                    "CompletedQty": "0",
                    "MouldNumber": base.get("MouldNumber", ""),
                    "Material": selected_catalog.get("MainMaterial", "") or base.get("Material", ""),
                    "MaterialLocation": selected_catalog.get("MaterialLocation", "") or base.get("MaterialLocation", ""),
                    "ColourMasterbatch": f"{selected_catalog.get('Colour', '')} / {selected_catalog.get('Additive', '')}".strip(" /"),
                    "MaterialWeightG": selected_catalog.get("MaterialWeightG", ""),
                    "SecondMaterialWeightG": selected_catalog.get("SecondMaterialWeightG", ""),
                    "MasterbatchWeightG": selected_catalog.get("MasterbatchWeightG", ""),
                    "UnitWeightG": selected_catalog.get("UnitWeightG", ""),
                    "Label": selected_catalog.get("HasLabel", ""),
                    "PackagingUnit": selected_catalog.get("PackagingUnit", ""),
                    "PackagingType": packaging_type_from_catalog(selected_catalog),
                    "PackagingOption": selected_catalog.get("PackagingOption", ""),
                    "CartonUnitStackQty": selected_catalog.get("CartonUnitStackQty", ""),
                    "PalletQty": selected_catalog.get("PalletQty", ""),
                    "AdditionalPackaging": selected_catalog.get("AdditionalPackaging", ""),
                    "PalletBag": selected_catalog.get("PalletBag", ""),
                    "PalletType": selected_catalog.get("PalletType", ""),
                    "WrapPallet": selected_catalog.get("WrapPallet", ""),
                    "FoodApplication": selected_catalog.get("FoodApplication", ""),
                    "CornerProtector": selected_catalog.get("CornerProtector", ""),
                    "InventoryLocationID": selected_catalog.get("InventoryLocationID", ""),
                    "AdditionalInstructions": selected_catalog.get("AdditionalInstructions", ""),
                    "QCRequired": "YES" if qc_required_for_save(qc_notes) else "NO",
                    "QCNotes": qc_notes_for_save(qc_notes),
                    "Notes": notes,
                },
            )
            set_flash(
                f"temp_plan_flash_{machine_id}",
                f"Production plan added: {schedule_id}. Queue sequence conflicts were adjusted automatically if needed.",
            )
            st.rerun()
        except ValueError as exc:
            st.error(str(exc))


def machine_schedule_management_panel(machine_id: str, machine_rows: pd.DataFrame) -> None:
    if not can_edit_production():
        return

    st.subheader("Correct / Delete Production Item")
    choices = [
        f"{row.ScheduleID} | #{row.Sequence or '-'} | {row.Status or '-'} | {row.ProductName or row.ProductCode or '(empty)'}"
        for row in machine_rows.itertuples()
    ]
    selected_choice = st.selectbox(
        "Production item",
        choices,
        key=f"machine_manage_schedule_{machine_id}",
    )
    selected_position = choices.index(selected_choice)
    row = machine_rows.iloc[selected_position].to_dict()
    schedule_id = str(row.get("ScheduleID", ""))
    all_status_options = ["Running", "Next", "Planned", "Paused", "Finished", "Idle", "Completed"]
    status_options = status_options_for_machine(machine_rows, machine_id, all_status_options, schedule_id)
    show_running_warning_if_needed(machine_rows, machine_id, schedule_id)
    current_status = row.get("Status", "")
    status_index = status_options.index(current_status) if current_status in status_options else 0

    move_cols = st.columns(2)
    with move_cols[0]:
        if st.button(
            "Move Up / 上移",
            key=f"move_up_{schedule_id}",
            disabled=selected_position == 0,
        ):
            try:
                move_production_item(current_user()["username"], schedule_id, "up")
                st.success("Queue order updated.")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))
    with move_cols[1]:
        if st.button(
            "Move Down / 下移",
            key=f"move_down_{schedule_id}",
            disabled=selected_position >= len(choices) - 1,
        ):
            try:
                move_production_item(current_user()["username"], schedule_id, "down")
                st.success("Queue order updated.")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))

    with st.expander("Edit selected item / 修正选中的生产项", expanded=False):
        with st.form(f"machine_edit_schedule_{schedule_id}"):
            sequence = st.number_input(
                "Sequence / 顺序",
                min_value=1,
                step=1,
                value=max(int_value(row.get("Sequence")), 1),
            )
            status = st.selectbox(
                "Status / 状态",
                status_options,
                index=status_index,
                format_func=production_status_label,
            )
            product_code = st.text_input("Product code / 产品编号", value=row.get("ProductCode", ""))
            product_name = st.text_input("Product name / 产品名称", value=row.get("ProductName", ""))
            planned = st.number_input(
                "Planned quantity / 计划数量",
                min_value=0,
                step=1,
                value=int_value(row.get("PlannedQty")),
            )
            completed = st.number_input(
                "Completed quantity / 完成数量",
                min_value=0,
                step=1,
                value=int_value(row.get("CompletedQty")),
            )
            qc_notes = st.text_area(
                t("production.qc_details"),
                value=str(row.get("QCNotes", "") or ""),
                height=80,
                placeholder=t("production.qc_details_placeholder"),
            )
            mould_number = st.text_input("Mould number / 模具编号", value=row.get("MouldNumber", ""))
            show_product_mould_hint(product_code, mould_number)
            render_production_mould_parameter_summary(product_code, machine_id, mould_number)
            override_reason = ""
            if can_force_incompatible_mould() and mould_machine_incompatibility(mould_number, machine_id):
                override_reason = st.text_input("Developer override reason / 开发者强制选择原因")
            material = st.text_input("Material / 材料", value=row.get("Material", ""))
            colour = st.text_input("Colour / masterbatch / 颜色", value=row.get("ColourMasterbatch", ""))
            notes = st.text_area("Notes / 备注", value=row.get("Notes", ""))
            save_submitted = st.form_submit_button("Save Correction / 保存修正")

        if save_submitted:
            values = row.copy()
            values.update(
                {
                    "Sequence": str(sequence),
                    "Status": status,
                    "ProductCode": product_code,
                    "ProductName": product_name,
                    "PlannedQty": str(planned),
                    "CompletedQty": str(completed),
                    "MouldNumber": mould_number,
                    "QCRequired": "YES" if qc_required_for_save(qc_notes) else "NO",
                    "QCNotes": qc_notes_for_save(qc_notes),
                    "Material": material,
                    "ColourMasterbatch": colour,
                    "Notes": notes,
                    "_MouldCompatibilityOverrideReason": override_reason,
                }
            )
            update_production_record(
                current_user()["username"],
                machine_id,
                values,
                schedule_id=schedule_id,
            )
            st.success("Production item corrected.")
            st.rerun()

    with st.expander("Delete selected item / 删除选中的生产项", expanded=False):
        st.warning("This removes the selected production item from ProductionSchedule.xlsx and records the deletion in ChangeLog.xlsx.")
        confirm_delete = st.checkbox(
            f"Confirm deletion of {schedule_id}",
            key=f"delete_confirm_{schedule_id}",
        )
        if st.button(
            "Delete Production Item",
            key=f"delete_schedule_{schedule_id}",
            disabled=not confirm_delete,
        ):
            delete_production_record(current_user()["username"], schedule_id)
            st.success("Production item deleted.")
            st.rerun()


def machine_page(machine_id: str = "", public_view: bool = False) -> None:
    archive_completed_for_machine_view(show_message=not public_view)
    production = get_production()
    machine_ids = registered_machine_ids(production)
    if not machine_id and machine_ids:
        machine_id = machine_ids[0]
    if not machine_id:
        st.warning("No machine records found.")
        return

    if len(machine_ids) > 1 and not public_view:
        machine_id = st.selectbox("Machine / 机器", machine_ids, index=machine_ids.index(machine_id) if machine_id in machine_ids else 0)

    machine_rows = production[production["MachineID"] == machine_id].copy()
    if machine_rows.empty:
        st.title(f"Machine {machine_id}")
        st.info("No production plan is assigned to this machine.")
        if not public_view:
            st.caption("Add the first plan from Production Table. This machine is already available in the machine selector.")
        return

    if public_view:
        running_rows = machine_rows[machine_rows["Status"].str.lower() == "running"].copy()
        if not running_rows.empty:
            machine_rows = running_rows

    machine_rows["_SortSequence"] = machine_rows["Sequence"].apply(int_value)
    machine_rows = machine_rows.sort_values(["_SortSequence", "ScheduleID"])
    first_row = machine_rows.iloc[0]
    running_for_alert = machine_rows[machine_rows["Status"].astype(str).str.strip().str.casefold().eq("running")]
    if not running_for_alert.empty:
        show_running_product_change_alert_once(running_for_alert.iloc[0], machine_id)
    total_planned = sum(int_value(value) for value in machine_rows["PlannedQty"].tolist())
    total_completed = sum(int_value(value) for value in machine_rows["CompletedQty"].tolist())
    total_remaining = max(total_planned - total_completed, 0)
    overall_progress = total_completed / total_planned * 100 if total_planned else 0

    title_prefix = "Running" if public_view and first_row.get("Status") == "Running" else str(first_row["MachineName"])
    st.title(f"{title_prefix} ({machine_id})")
    st.markdown(
        f"""
        <div class="quick-summary">
        <div class="metric-grid">
            <div class="metric-card"><div class="metric-label">Items</div><div class="metric-value">{len(machine_rows)}</div></div>
            <div class="metric-card"><div class="metric-label">Total Planned</div><div class="metric-value">{total_planned:,}</div></div>
            <div class="metric-card"><div class="metric-label">Total Completed</div><div class="metric-value">{total_completed:,}</div></div>
            <div class="metric-card"><div class="metric-label">Total Remaining</div><div class="metric-value">{total_remaining:,}</div></div>
        </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.progress(min(overall_progress / 100, 1.0) if total_planned else 0)
    st.caption(f"Overall progress: {overall_progress:.1f}%")

    for position, (_, row) in enumerate(machine_rows.iterrows(), start=1):
        st.markdown(product_schedule_card(row, position), unsafe_allow_html=True)
        if not public_view and can_view_mould_parameters():
            render_production_mould_parameter_summary(row.get("ProductCode", ""), machine_id, row.get("MouldNumber", ""), compact=True)

    if not public_view:
        machine_schedule_management_panel(machine_id, machine_rows.reset_index(drop=True))
        add_temp_plan_panel(machine_id, machine_rows)


def stock_in_page(public_mode: bool = False) -> None:
    if not stock_in_access_allowed(public_mode):
        return
    if current_user() and not can_stock_in():
        st.error("Your account does not have permission to perform stock-in.")
        return
    st.title("Warehouse Stock-In")
    show_flash("stock_in_flash")
    st.markdown('<div class="stock-in-route"></div>', unsafe_allow_html=True)
    inventory = get_inventory()
    catalog = get_product_catalog()
    search_mode = st.radio(
        "Stock-in source / 入库来源",
        ["Running Machine / 当前机器生产", "Inventory / 库存表", "Product Catalog / 产品目录"],
        horizontal=False,
    )

    product_code = ""
    product_name = ""
    machine_id = ""
    schedule_id = ""
    production_status = ""
    mould_number = ""
    unit = "pcs"
    location = ""
    current_stock = "0"
    create_if_missing = False

    if search_mode == "Running Machine / 当前机器生产":
        selected_running = running_machine_product_selector()
        if selected_running is None:
            return
        machine_id = selected_running["machine_id"]
        schedule_id = selected_running.get("schedule_id", "")
        production_status = selected_running.get("production_status", "")
        product_code = selected_running["product_code"]
        product_name = selected_running["product_name"]
        mould_number = selected_running["mould_number"]
        unit = selected_running["unit"]
        location = selected_running["location"]
        current_stock = selected_running["current_stock"]
        create_if_missing = selected_running["create_if_missing"]
        if create_if_missing:
            st.info("This running product is not in Inventory.xlsx yet. Stock-in will create it automatically.")
    elif search_mode == "Inventory / 库存表":
        keyword = st.text_input("Search inventory / 搜索库存", placeholder="Product code, name, location")
        filtered = inventory
        if keyword:
            words = [word.strip() for word in keyword.replace(",", " ").split() if word.strip()]
            for word in words:
                mask = (
                    filtered["ProductCode"].str.contains(word, case=False, na=False, regex=False)
                    | filtered["ProductName"].str.contains(word, case=False, na=False, regex=False)
                    | filtered["Location"].str.contains(word, case=False, na=False, regex=False)
                )
                filtered = filtered[mask]

        if filtered.empty:
            st.warning("No matching inventory products.")
            return

        options = [f"{r.ProductCode} - {r.ProductName} (Stock: {r.CurrentStock} {r.Unit}, {r.Location or '-'})" for r in filtered.itertuples()]
        selected = st.selectbox("Product", options)
        product_code = selected.split(" - ", 1)[0]
        row = inventory[inventory["ProductCode"] == product_code].iloc[0]
        product_name = row["ProductName"]
        unit = row["Unit"] or "pcs"
        location = row["Location"]
        current_stock = row["CurrentStock"]
    else:
        if catalog.empty:
            st.warning("No product catalog found.")
            return
        filtered_catalog = apply_catalog_search_filters(catalog, "stock_catalog")
        st.write(f"Matched catalog products: **{len(filtered_catalog)}**")
        if filtered_catalog.empty:
            st.warning("No matching catalog products.")
            return
        filtered_catalog = filtered_catalog.reset_index(drop=True)
        choices = [product_choice_label(row, i) for i, row in enumerate(filtered_catalog.itertuples())]
        selected = st.selectbox("Product", choices)
        selected_catalog = filtered_catalog.iloc[choices.index(selected)].to_dict()
        product_code = selected_catalog.get("Item", "")
        product_name = selected_catalog.get("ProductDetail", "") or product_code
        inventory_match = inventory[inventory["ProductCode"] == product_code]
        if not inventory_match.empty:
            row = inventory_match.iloc[0]
            unit = row["Unit"] or "pcs"
            location = row["Location"]
            current_stock = row["CurrentStock"]
        else:
            create_if_missing = True
            location = selected_catalog.get("InventoryLocationID", "")
            st.info("This product is not in Inventory.xlsx yet. Stock-in will create it automatically.")

    st.markdown(
        f"""
        <div class="info-card stock-selected-card">
            <div class="stock-selected-title">{product_code} - {product_name}</div>
            <div class="machine-meta-grid">
                <div class="machine-meta">
                    <div class="machine-meta-label">Current stock</div>
                    <div class="machine-meta-value">{current_stock} {unit}</div>
                </div>
                <div class="machine-meta">
                    <div class="machine-meta-label">Location</div>
                    <div class="machine-meta-value">{location or "-"}</div>
                </div>
            </div>
            {compact_field("Machine", machine_id)}
            {compact_field("Mould", mould_number)}
        </div>
        """,
        unsafe_allow_html=True,
    )

    if "local_stock_in_request_id" not in st.session_state:
        st.session_state["local_stock_in_request_id"] = f"local-stock-in-{uuid4()}"
    local_request_id = st.session_state["local_stock_in_request_id"]

    with st.form("stock_in_form"):
        qty = st.number_input("Stock-in quantity", min_value=1, step=1)
        unit = st.text_input("Unit", value=unit or "pcs")
        location = st.text_input("Inventory location", value=location)
        default_remarks = f"Stock-in from {machine_id}, mould {mould_number}" if machine_id else ""
        remarks = st.text_input("Remarks", value=default_remarks)
        st.caption(f"Request ID: {local_request_id}")
        submitted = st.form_submit_button("Submit Stock-In")
    if submitted:
        try:
            user_name = current_user()["username"] if current_user() else "stock-in-qr"
            old_stock, new_stock = stock_in(
                user_name,
                product_code,
                int(qty),
                remarks,
                product_name=product_name,
                unit=unit,
                location=location,
                create_if_missing=create_if_missing,
                machine_id=machine_id,
                schedule_id=schedule_id,
                mould_number=mould_number,
                production_status=production_status,
                client_request_id=local_request_id,
            )
            set_flash(
                "stock_in_flash",
                f"Stock-In processed once. Request ID: {local_request_id}. Inventory updated: {old_stock:,} -> {new_stock:,}.",
            )
            st.session_state["local_stock_in_request_id"] = f"local-stock-in-{uuid4()}"
            st.rerun()
        except ValueError as exc:
            st.error(str(exc))

    st.info("Photo upload is disabled on the local Stock-In page because image files are not saved with the Excel stock-in record.")


def loose_goods_page() -> None:
    if not can_manage_loose_goods():
        st.error("Your account does not have permission to manage loose goods.")
        return

    st.title(f"{t('loose.title')} / 散货")
    loose = get_loose_goods()
    production = get_production()

    st.subheader("Create from production / 从生产表创建")
    active = production[production["Status"].astype(str).isin(["Running", "Next", "Queued"])].copy()
    if active.empty:
        st.info("No Running or Queued production items are available.")
    else:
        active["_SequenceSort"] = pd.to_numeric(active["Sequence"], errors="coerce").fillna(999999)
        active = active.sort_values(["MachineID", "_SequenceSort"]).reset_index(drop=True)
        choices = [
            f"{row.ScheduleID} | {row.MachineID} | {row.Status} | {row.ProductCode} | {row.ProductName}"
            for row in active.itertuples()
        ]
        with st.form("create_loose_goods_form"):
            selected = st.selectbox("Production item / 生产项目", choices)
            quantity = st.number_input(f"{t('loose.quantity')} / 数量", min_value=1, step=1)
            status = st.selectbox(
                f"{t('loose.status')} / 状态",
                ["Loose", "WaitingForWrap", "WaitingForHandle", "ReadyForStockIn"],
                format_func=lambda value: {
                    "Loose": "Loose / 散货",
                    "WaitingForWrap": "Waiting for Wrap / 等待缠绕",
                    "WaitingForHandle": "Waiting for Handle / 等安装把手",
                    "ReadyForStockIn": "Ready for Stock-In / 可正式入库",
                }.get(value, value),
            )
            operator = st.text_input("Operator / 操作员", value=current_user()["username"])
            notes = st.text_area("Notes / 备注")
            submitted = st.form_submit_button(f"{t('loose.create')} / 创建")
        if submitted:
            row = active.iloc[choices.index(selected)].to_dict()
            try:
                loose_id = add_loose_goods_record(
                    user=current_user()["username"],
                    machine_id=str(row.get("MachineID", "")),
                    schedule_id=str(row.get("ScheduleID", "")),
                    product_code=str(row.get("ProductCode", "")),
                    product_name=str(row.get("ProductName", "")),
                    mould_number=str(row.get("MouldNumber", "")),
                    quantity=int(quantity),
                    status=status,
                    operator=operator,
                    notes=notes,
                    request_type="admin_loose",
                )
                st.success(f"Loose goods record created: {loose_id}")
                st.rerun()
            except Exception as exc:
                st.error(str(exc))

    st.subheader("Manage loose goods / 管理散货")
    loose = get_loose_goods()
    if loose.empty:
        st.info("No loose goods records yet.")
        return

    filter_cols = st.columns(3)
    with filter_cols[0]:
        machine_filter = st.selectbox("Machine", option_values(loose, "MachineID"), key="loose_machine_filter")
    with filter_cols[1]:
        product_filter = st.text_input("Product / 产品", key="loose_product_filter")
    with filter_cols[2]:
        status_filter = st.selectbox("Status", option_values(loose, "Status"), key="loose_status_filter")

    filtered = loose.copy()
    if machine_filter != "Any":
        filtered = filtered[filtered["MachineID"].astype(str) == machine_filter]
    if status_filter != "Any":
        filtered = filtered[filtered["Status"].astype(str) == status_filter]
    if product_filter:
        mask = (
            filtered["ProductCode"].astype(str).str.contains(product_filter, case=False, na=False)
            | filtered["ProductName"].astype(str).str.contains(product_filter, case=False, na=False)
        )
        filtered = filtered[mask]

    st.dataframe(filtered.sort_index(ascending=False), use_container_width=True, hide_index=True)

    editable = filtered[filtered["Status"].astype(str) != "StockedIn"].copy()
    if not editable.empty:
        st.subheader("Update status / 更新状态")
        labels = [f"{row.LooseID} | {row.MachineID} | {row.ProductCode} | {row.Quantity} | {row.Status}" for row in editable.itertuples()]
        with st.form("update_loose_status_form"):
            selected = st.selectbox("Loose goods record", labels)
            new_status = st.selectbox(
                "New status",
                ["Loose", "WaitingForWrap", "WaitingForHandle", "ReadyForStockIn"],
                format_func=lambda value: {
                    "Loose": "Loose / 散货",
                    "WaitingForWrap": "Waiting for Wrap / 等待缠绕",
                    "WaitingForHandle": "Waiting for Handle / 等安装把手",
                    "ReadyForStockIn": "Ready for Stock-In / 可正式入库",
                }.get(value, value),
            )
            update_notes = st.text_input("Update notes / 更新备注")
            save_status = st.form_submit_button("Save status / 保存状态")
        if save_status:
            loose_id = str(editable.iloc[labels.index(selected)]["LooseID"])
            try:
                update_loose_goods_status(current_user()["username"], loose_id, new_status, update_notes)
                st.success(f"Updated {loose_id}")
                st.rerun()
            except Exception as exc:
                st.error(str(exc))

    ready = filtered[filtered["Status"].astype(str) == "ReadyForStockIn"].copy()
    if not ready.empty:
        st.subheader("Convert to Stock-In / 转正式入库")
        ready_labels = [f"{row.LooseID} | {row.MachineID} | {row.ProductCode} | {row.Quantity}" for row in ready.itertuples()]
        selected_ready = st.selectbox("Ready loose goods", ready_labels)
        convert_notes = st.text_input("Stock-in remarks / 入库备注", key="loose_convert_notes")
        if st.button("Convert selected loose goods to stock-in / 转正式入库", type="primary"):
            loose_id = str(ready.iloc[ready_labels.index(selected_ready)]["LooseID"])
            try:
                old_stock, new_stock = stock_in_from_loose_goods(current_user()["username"], loose_id, convert_notes)
                st.success(f"Converted {loose_id}: stock {old_stock:,} -> {new_stock:,}")
                st.rerun()
            except Exception as exc:
                st.error(str(exc))


PRODUCT_TYPE_BUCKET = "\u6876"
PRODUCT_TYPE_LID = "\u76d6\u5b50"
PRODUCT_TYPE_MANUAL = "__manual_search__"
FILTER_ANY = "Any"
FILTER_YES = "Yes"
FILTER_NO = "No"


def option_values(df: pd.DataFrame, column: str) -> list[str]:
    if df.empty or column not in df.columns:
        return [FILTER_ANY]
    values = sorted({str(value).strip() for value in df[column].dropna().tolist() if str(value).strip()})
    return [FILTER_ANY] + values


def filter_option_label(value: object) -> str:
    text = str(value or "")
    return t("common.any") if text == FILTER_ANY else text


def yes_no_filter_label(value: object) -> str:
    text = str(value or "")
    return {
        FILTER_ANY: t("common.any"),
        FILTER_YES: t("common.yes"),
        FILTER_NO: t("common.no"),
    }.get(text, text)


def product_type_filter_label(value: object) -> str:
    return {
        PRODUCT_TYPE_BUCKET: t("production.type.bucket"),
        PRODUCT_TYPE_LID: t("production.type.lid"),
        PRODUCT_TYPE_MANUAL: t("production.type.manual_search"),
    }.get(str(value or ""), str(value or ""))


def apply_catalog_search_filters(catalog: pd.DataFrame, key_prefix: str, default_type: str = PRODUCT_TYPE_BUCKET) -> pd.DataFrame:
    filtered = catalog.copy()
    type_options = [PRODUCT_TYPE_BUCKET, PRODUCT_TYPE_LID, PRODUCT_TYPE_MANUAL]
    default_choice = default_type if default_type in type_options else PRODUCT_TYPE_BUCKET

    product_type = st.segmented_control(
        t("production.product_type"),
        type_options,
        default=default_choice,
        key=f"{key_prefix}_type_i18n",
        format_func=product_type_filter_label,
    )
    if product_type in [PRODUCT_TYPE_BUCKET, PRODUCT_TYPE_LID]:
        filtered = filtered[filtered["ProductType"] == product_type]

    keyword = st.text_input(
        t("production.keywords"),
        placeholder=t("production.keywords_placeholder"),
        key=f"{key_prefix}_keyword",
    )
    if keyword:
        words = [word.strip() for word in keyword.replace(",", " ").split() if word.strip()]
        search_columns = [
            "ProductDetail",
            "Item",
            "Part",
            "Size",
            "Colour",
            "HasLabel",
            "PackagingUnit",
            "AdditionalPackaging",
            "PackagingType",
            "CycleTime",
            "ShotWeight",
            "Cavity",
            "MainMaterial",
            "Additive",
            "AdditionalInstructions",
        ]
        for word in words:
            mask = pd.Series(False, index=filtered.index)
            for column in search_columns:
                mask = mask | filtered[column].astype(str).str.contains(word, case=False, na=False, regex=False)
            filtered = filtered[mask]

    size = st.selectbox(t("production.size"), option_values(filtered, "Size"), key=f"{key_prefix}_size", format_func=filter_option_label)
    if size != FILTER_ANY:
        filtered = filtered[filtered["Size"] == size]

    colour = st.selectbox(t("production.colour"), option_values(filtered, "Colour"), key=f"{key_prefix}_colour", format_func=filter_option_label)
    if colour != FILTER_ANY:
        filtered = filtered[filtered["Colour"] == colour]

    label_choice = st.selectbox(t("production.label"), [FILTER_ANY, FILTER_YES, FILTER_NO], key=f"{key_prefix}_label", format_func=yes_no_filter_label)
    if label_choice == FILTER_YES:
        filtered = filtered[filtered["HasLabel"].apply(lambda value: is_meaningful(value) and str(value).strip().casefold() != "no")]
    elif label_choice == FILTER_NO:
        filtered = filtered[~filtered["HasLabel"].apply(lambda value: is_meaningful(value) and str(value).strip().casefold() != "no")]

    with st.expander(t("production.more_filters"), expanded=False):
        material = st.selectbox(t("production.material"), option_values(filtered, "MainMaterial"), key=f"{key_prefix}_material", format_func=filter_option_label)
        if material != FILTER_ANY:
            filtered = filtered[filtered["MainMaterial"] == material]

        packaging = st.selectbox(t("production.packaging"), option_values(filtered, "PackagingUnit"), key=f"{key_prefix}_packaging", format_func=filter_option_label)
        if packaging != FILTER_ANY:
            filtered = filtered[filtered["PackagingUnit"] == packaging]

        special_only = st.checkbox(t("production.special_only"), key=f"{key_prefix}_special")
        if special_only:
            special_columns = ["AdditionalPackaging", "PalletBag", "WrapPallet", "FoodApplication", "CornerProtector", "AdditionalInstructions"]
            mask = pd.Series(False, index=filtered.index)
            for column in special_columns:
                mask = mask | filtered[column].apply(is_meaningful)
            filtered = filtered[mask]

    return filtered


def product_choice_label(row: object, index: int) -> str:
    option = packaging_option_text(row)
    option_text = f" | {t('production.packaging_option')}: {option}" if option else ""
    return f"{index + 1}. {row.ProductDetail or row.Item}{option_text} | {row.Size or '-'} | {row.Colour or '-'} | {t('production.label_short')}: {row.HasLabel or '-'}"


def forecast_review_state(review_due_date: object) -> tuple[str, str]:
    parsed = pd.to_datetime(review_due_date, errors="coerce")
    if pd.isna(parsed):
        return "forecast-red", t("production.forecast_overdue")
    days_left = (parsed.normalize() - pd.Timestamp.today().normalize()).days
    if days_left < 0:
        return "forecast-red", t("production.forecast_overdue_days", days=abs(days_left))
    if days_left <= 7:
        return "forecast-yellow", t("production.forecast_due_days", days=days_left)
    return "forecast-green", t("production.forecast_recent")


def forecast_widget_key(prefix: str, product_code: object, product_name: object = "") -> str:
    raw = str(product_code or product_name or "unknown")
    safe = "".join(ch if ch.isalnum() else "_" for ch in raw)[:90] or "unknown"
    return f"{prefix}_{safe}"


def jump_to_forecast_product(product_code: object) -> None:
    code = str(product_code or "").strip()
    st.session_state["active_page"] = "forecast"
    st.session_state["_last_requested_page"] = "forecast"
    st.query_params["page"] = "forecast"
    if code:
        st.query_params["product_code"] = code
    else:
        try:
            del st.query_params["product_code"]
        except Exception:
            pass
    st.rerun()


def show_latest_forecast_for_product(product_code: object, product_name: object) -> None:
    code = str(product_code or "").strip()
    name = str(product_name or "").strip()
    st.subheader(t("production.forecast"))
    result = get_latest_forecast_result(product_code=code, product_name=name)
    if not result:
        st.markdown(
            f"""
            <div class="forecast-summary-card forecast-red">
                <span class="forecast-badge forecast-red">{escape(t("production.no_forecast"))}</span>
                <div class="field-label">{escape(t("production.product_code"))}</div>
                <div class="field-value">{escape(code or "-")}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button(
            t("production.run_forecast"),
            key=forecast_widget_key("forecast_jump_empty", code, name),
        ):
            jump_to_forecast_product(code)
        return

    status_css, status_text = forecast_review_state(result.get("review_due_date"))
    weights = " / ".join(
        [
            f"A {float(result.get('best_weight_A') or 0):.1f}",
            f"B {float(result.get('best_weight_B') or 0):.1f}",
            f"C {float(result.get('best_weight_C') or 0):.1f}",
        ]
    )
    st.markdown(
        f"""
        <div class="forecast-summary-card {status_css}">
            <span class="forecast-badge {status_css}">{escape(status_text)}</span>
            <div class="machine-meta-grid">
                <div class="machine-meta"><div class="machine-meta-label">{escape(t("production.recommended_qty"))}</div><div class="machine-meta-value">{int_value(result.get('recommended_production_quantity')):,}</div></div>
                <div class="machine-meta"><div class="machine-meta-label">{escape(t("production.next_1_month"))}</div><div class="machine-meta-value">{int_value(result.get('forecast_next_1_month')):,}</div></div>
                <div class="machine-meta"><div class="machine-meta-label">{escape(t("production.next_2_months"))}</div><div class="machine-meta-value">{int_value(result.get('forecast_next_2_months')):,}</div></div>
                <div class="machine-meta"><div class="machine-meta-label">{escape(t("production.safety_stock"))}</div><div class="machine-meta-value">{int_value(result.get('safety_stock')):,}</div></div>
            </div>
            {compact_field(t("production.weights"), weights, show_no=True)}
            {compact_field(t("production.mae_mape"), f"{result.get('mae', 0)} / {result.get('mape', 0)}", show_no=True)}
            {compact_field(t("production.run_date"), result.get("run_date") or result.get("timestamp"), show_no=True)}
            {compact_field(t("production.review_due"), result.get("review_due_date"), show_no=True)}
            {compact_field(t("production.notes"), result.get("notes"), show_no=True)}
        </div>
        """,
        unsafe_allow_html=True,
    )
    if st.button(
        t("production.run_forecast"),
        key=forecast_widget_key("forecast_jump_existing", code, name),
    ):
        jump_to_forecast_product(code)


def active_mould_for_product(product_code: object) -> dict[str, str]:
    primary = get_primary_mould(str(product_code or ""))
    if not primary:
        return {"MouldNumber": "", "MouldName": ""}
    return {
        "MouldNumber": str(primary.get("MouldNumber", "") or ""),
        "MouldName": str(primary.get("MouldName", "") or primary.get("ProductName", "") or ""),
    }


def mould_machine_incompatibility(mould_number: object, machine_id: object) -> bool:
    mould_number = str(mould_number or "").strip()
    machine_id = str(machine_id or "").strip()
    if not mould_number or not machine_id:
        return False
    compatible = get_mould_compatible_machine_ids(mould_number)
    return bool(compatible and machine_id not in compatible)


def _value_or_dash(value: object) -> str:
    text = str(value or "").strip()
    return text if text else "-"


def render_small_table(title: str, rows: list[dict[str, object]]) -> None:
    st.markdown(f"**{title}**")
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def render_mould_setting_view(bundle: dict[str, object], full: bool = True) -> None:
    setting = bundle.get("setting") if isinstance(bundle, dict) else None
    if not setting:
        st.info("No parameter configuration yet / 尚未建立参数配置。")
        return
    setting = dict(setting)
    st.caption(
        f"Version V{_value_or_dash(setting.get('Version'))} | "
        f"Updated: {_value_or_dash(setting.get('UpdatedAt'))} | "
        f"By: {_value_or_dash(setting.get('UpdatedBy'))}"
    )

    render_small_table(
        "Injection 1-5 / 注塑1-5段",
        [
            {
                "Stage": f"Injection {stage}",
                "Pressure": _value_or_dash(setting.get(f"injection_stage_{stage}_pressure")),
                "Speed": _value_or_dash(setting.get(f"injection_stage_{stage}_speed")),
                "Position": _value_or_dash(setting.get(f"injection_stage_{stage}_position")),
            }
            for stage in range(1, 6)
        ],
    )
    render_small_table(
        "Holding / 保压",
        [
            {
                "Stage": f"Holding {stage}",
                "Pressure": _value_or_dash(setting.get(f"holding_stage_{stage}_pressure")),
                "Speed": _value_or_dash(setting.get(f"holding_stage_{stage}_speed")),
                "Time": _value_or_dash(setting.get(f"holding_stage_{stage}_time")),
            }
            for stage in range(1, 4)
        ],
    )
    render_small_table(
        "Core time / 核心时间",
        [
            {"Parameter": "Injection time (s)", "Value": _value_or_dash(setting.get("injection_time_seconds"))},
            {"Parameter": "Holding time (s)", "Value": _value_or_dash(setting.get("holding_time_seconds"))},
            {"Parameter": "Cooling time (s)", "Value": _value_or_dash(setting.get("cooling_time_seconds"))},
            {"Parameter": "Cycle time (s)", "Value": _value_or_dash(setting.get("cycle_time_seconds"))},
        ],
    )
    render_small_table(
        "Barrel temperature / 炮筒温度",
        [
            {
                "Nozzle C": _value_or_dash(setting.get("nozzle_temperature")),
                "Barrel 1 C": _value_or_dash(setting.get("barrel_temperature_1")),
                "Barrel 2 C": _value_or_dash(setting.get("barrel_temperature_2")),
                "Barrel 3 C": _value_or_dash(setting.get("barrel_temperature_3")),
                "Barrel 4 C": _value_or_dash(setting.get("barrel_temperature_4")),
            }
        ],
    )

    hot = bundle.get("hot_runner")
    if isinstance(hot, pd.DataFrame) and not hot.empty:
        show_cols = [col for col in ["ZoneNumber", "ZoneName", "Temperature"] if col in hot.columns]
        st.markdown("**Hot runner / 热流道温度**")
        st.dataframe(hot[show_cols], use_container_width=True, hide_index=True)

    if not full:
        return

    render_small_table(
        "Plasticizing / 溶胶塑化",
        [
            {"Parameter": "Plasticizing pressure", "Value": _value_or_dash(setting.get("plasticizing_pressure"))},
            {"Parameter": "Plasticizing speed", "Value": _value_or_dash(setting.get("plasticizing_speed"))},
            {"Parameter": "Plasticizing quantity", "Value": _value_or_dash(setting.get("plasticizing_quantity"))},
            {"Parameter": "Back pressure", "Value": _value_or_dash(setting.get("back_pressure"))},
            {"Parameter": "Decompression quantity", "Value": _value_or_dash(setting.get("decompression_quantity"))},
        ],
    )
    render_small_table(
        "Ejector / 顶针",
        [
            {
                "Action": "Eject out 1 / 顶出一段",
                "Pressure": _value_or_dash(setting.get("ejector_stage_1_pressure")),
                "Speed": _value_or_dash(setting.get("ejector_stage_1_speed")),
                "Distance": _value_or_dash(setting.get("ejector_stage_1_distance")),
                "Delay": _value_or_dash(setting.get("ejector_stage_1_delay")),
            },
            {
                "Action": "Eject out 2 / 顶出二段",
                "Pressure": _value_or_dash(setting.get("ejector_stage_2_pressure")),
                "Speed": _value_or_dash(setting.get("ejector_stage_2_speed")),
                "Distance": _value_or_dash(setting.get("ejector_stage_2_distance")),
                "Delay": _value_or_dash(setting.get("ejector_stage_2_delay")),
            },
            {
                "Action": "Return 1 / 返回一段",
                "Pressure": _value_or_dash(setting.get("ejector_return_stage_1_pressure")),
                "Speed": _value_or_dash(setting.get("ejector_return_stage_1_speed")),
                "Distance": _value_or_dash(setting.get("ejector_return_stage_1_distance")),
                "Delay": _value_or_dash(setting.get("ejector_return_stage_1_delay")),
            },
            {
                "Action": "Return 2 / 返回二段",
                "Pressure": _value_or_dash(setting.get("ejector_return_stage_2_pressure")),
                "Speed": _value_or_dash(setting.get("ejector_return_stage_2_speed")),
                "Distance": _value_or_dash(setting.get("ejector_return_stage_2_distance")),
                "Delay": _value_or_dash(setting.get("ejector_return_stage_2_delay")),
            },
        ],
    )
    air_rows = [
        {
            "Group": f"Moving air {stage}",
            "Time": _value_or_dash(setting.get(f"moving_air_{stage}_time")),
            "Delay": _value_or_dash(setting.get(f"moving_air_{stage}_delay")),
            "Position": _value_or_dash(setting.get(f"moving_air_{stage}_position")),
        }
        for stage in range(1, 4)
    ]
    air_rows.append(
        {
            "Group": "Fixed air 1",
            "Time": _value_or_dash(setting.get("fixed_air_1_time")),
            "Delay": _value_or_dash(setting.get("fixed_air_1_delay")),
            "Position": _value_or_dash(setting.get("fixed_air_1_position")),
        }
    )
    render_small_table("Air blow / 吹气", air_rows)

    robot = bundle.get("robot")
    if isinstance(robot, pd.DataFrame) and not robot.empty:
        show_cols = [col for col in ["ParameterName", "ParameterValue"] if col in robot.columns]
        st.markdown("**Robot important parameters / 机械手重要参数**")
        st.dataframe(robot[show_cols], use_container_width=True, hide_index=True)
    if is_meaningful(setting.get("Notes"), show_no=True):
        st.markdown("**Parameter notes / 参数备注**")
        st.write(setting.get("Notes"))


def render_production_mould_parameter_summary(product_code: object, machine_id: object, mould_number: object = "", compact: bool = False) -> None:
    product_code = str(product_code or "").strip()
    machine_id = str(machine_id or "").strip()
    mould_number = str(mould_number or "").strip()
    if not mould_number and product_code:
        mould_number = active_mould_for_product(product_code).get("MouldNumber", "")
    if not mould_number:
        st.info(t("moulds.no_default_link"))
        return
    moulds = get_moulds()
    match = moulds[moulds["MouldNumber"].astype(str).str.casefold().eq(mould_number.casefold())]
    mould_name = str(match.iloc[0].get("MouldName", "") or match.iloc[0].get("AssociatedProduct", "") or "") if not match.empty else ""
    compatible = get_mould_compatible_machine_ids(mould_number)
    compat_text = ", ".join(localized_machine_text(m) for m in compatible) if compatible else t("common.not_set")
    st.markdown(
        f"**{t('common.mould')}:** {mould_number} {mould_name or ''}  \\n"
        f"**{t('moulds.compatible_machines')}:** {compat_text}  \\n"
        f"**{t('moulds.current_machine')}:** {localized_machine_text(machine_id)}"
    )
    if machine_id and compatible and machine_id not in compatible:
        st.error(t("moulds.machine_not_compatible", mould=mould_number, machine=localized_machine_text(machine_id)))
    if not machine_id:
        return
    bundle = get_mould_machine_parameter_bundle(mould_number, machine_id)
    setting = bundle.get("setting")
    if not setting:
        st.warning(t("moulds.parameter_missing_for_machine", mould=mould_number, name=mould_name or "-", machine=localized_machine_text(machine_id)))
        if can_edit_mould_parameters():
            st.caption(t("moulds.open_moulds_to_create_parameters"))
        return
    st.caption(t("moulds.parameter_version_updated", version=setting.get("Version"), updated=setting.get("UpdatedAt") or "-"))
    with st.expander(t("moulds.show_core_parameters"), expanded=False):
        render_mould_setting_view(bundle, full=False)
    if not compact:
        with st.expander(t("moulds.view_full_parameters"), expanded=False):
            render_mould_setting_view(bundle, full=True)


def _setting_input(current: dict[str, object], field: str, label: str, key_prefix: str) -> str:
    return st.text_input(label, value=str(current.get(field, "") or ""), key=f"{key_prefix}_{field}")


def collect_mould_setting_form(current: dict[str, object], key_prefix: str, current_hot: pd.DataFrame | None = None, current_robot: pd.DataFrame | None = None) -> tuple[dict[str, str], list[dict[str, str]], list[dict[str, str]], str]:
    values: dict[str, str] = {}
    st.markdown("### Production core parameters / 生产核心参数")
    st.markdown("**Injection 1-5 / 注塑1-5段**")
    for stage in range(1, 6):
        cols = st.columns([0.9, 1, 1, 1])
        cols[0].markdown(f"Injection {stage}")
        with cols[1]:
            values[f"injection_stage_{stage}_pressure"] = _setting_input(current, f"injection_stage_{stage}_pressure", "Pressure", key_prefix)
        with cols[2]:
            values[f"injection_stage_{stage}_speed"] = _setting_input(current, f"injection_stage_{stage}_speed", "Speed", key_prefix)
        with cols[3]:
            values[f"injection_stage_{stage}_position"] = _setting_input(current, f"injection_stage_{stage}_position", "Position", key_prefix)

    st.markdown("**Holding / 保压**")
    for stage in range(1, 4):
        cols = st.columns([0.9, 1, 1, 1])
        cols[0].markdown(f"Holding {stage}")
        with cols[1]:
            values[f"holding_stage_{stage}_pressure"] = _setting_input(current, f"holding_stage_{stage}_pressure", "Pressure", key_prefix)
        with cols[2]:
            values[f"holding_stage_{stage}_speed"] = _setting_input(current, f"holding_stage_{stage}_speed", "Speed", key_prefix)
        with cols[3]:
            values[f"holding_stage_{stage}_time"] = _setting_input(current, f"holding_stage_{stage}_time", "Time", key_prefix)

    cols = st.columns(4)
    with cols[0]:
        values["injection_time_seconds"] = _setting_input(current, "injection_time_seconds", "Injection time (s)", key_prefix)
    with cols[1]:
        values["holding_time_seconds"] = _setting_input(current, "holding_time_seconds", "Holding total (s)", key_prefix)
    with cols[2]:
        values["cooling_time_seconds"] = _setting_input(current, "cooling_time_seconds", "Cooling time (s)", key_prefix)
    with cols[3]:
        values["cycle_time_seconds"] = _setting_input(current, "cycle_time_seconds", "Cycle time (s)", key_prefix)

    st.markdown("### Plasticizing and temperature / 塑化与温度")
    left, right = st.columns(2)
    with left:
        for field, label in [
            ("plasticizing_pressure", "Plasticizing pressure"),
            ("plasticizing_speed", "Plasticizing speed"),
            ("plasticizing_quantity", "Plasticizing quantity"),
            ("back_pressure", "Back pressure"),
            ("decompression_quantity", "Decompression quantity"),
        ]:
            values[field] = _setting_input(current, field, label, key_prefix)
    with right:
        temp_cols = st.columns(5)
        for column, (field, label) in zip(
            temp_cols,
            [
                ("nozzle_temperature", "Nozzle C"),
                ("barrel_temperature_1", "Barrel 1 C"),
                ("barrel_temperature_2", "Barrel 2 C"),
                ("barrel_temperature_3", "Barrel 3 C"),
                ("barrel_temperature_4", "Barrel 4 C"),
            ],
        ):
            with column:
                values[field] = _setting_input(current, field, label, key_prefix)

    st.markdown("### Ejector and air blow / 顶针与吹气")
    for title, prefix in [("Eject out / 顶出", "ejector_stage"), ("Return / 返回", "ejector_return_stage")]:
        st.markdown(f"**{title}**")
        for stage, stage_label in [(1, "Stage 1 / 一段"), (2, "Stage 2 / 二段")]:
            cols = st.columns([0.9, 1, 1, 1, 1])
            cols[0].markdown(stage_label)
            for idx, name in enumerate(["pressure", "speed", "distance", "delay"], start=1):
                field = f"{prefix}_{stage}_{name}"
                with cols[idx]:
                    values[field] = _setting_input(current, field, name.title(), key_prefix)

    for stage in range(1, 4):
        cols = st.columns([0.9, 1, 1, 1])
        cols[0].markdown(f"Moving air {stage}")
        for idx, name in enumerate(["time", "delay", "position"], start=1):
            field = f"moving_air_{stage}_{name}"
            with cols[idx]:
                values[field] = _setting_input(current, field, name.title(), key_prefix)
    cols = st.columns([0.9, 1, 1, 1])
    cols[0].markdown("Fixed air 1")
    for idx, field in enumerate(["fixed_air_1_time", "fixed_air_1_delay", "fixed_air_1_position"], start=1):
        with cols[idx]:
            values[field] = _setting_input(current, field, field.replace("fixed_air_1_", "").title(), key_prefix)

    st.markdown("### Hot runner / 热流道")
    hot_lookup: dict[str, dict[str, object]] = {}
    if isinstance(current_hot, pd.DataFrame) and not current_hot.empty:
        for _, hot_row in current_hot.iterrows():
            key = str(hot_row.get("ZoneNumber", "") or hot_row.get("DisplayOrder", "")).strip()
            if key:
                hot_lookup[key] = hot_row.to_dict()
    hot_rows: list[dict[str, str]] = []
    for order, zone_number in enumerate(["1", "2", "3", "4"], start=1):
        existing_hot = hot_lookup.get(zone_number, {})
        zone_label = f"Zone {zone_number}"
        cols = st.columns([1, 1.4, 1])
        cols[0].markdown(zone_label)
        with cols[1]:
            zone_name = st.text_input("Zone name", value=str(existing_hot.get("ZoneName", "") or zone_label), key=f"{key_prefix}_hot_name_{zone_number}")
        with cols[2]:
            temp = st.text_input("Temperature C", value=str(existing_hot.get("Temperature", "") or ""), key=f"{key_prefix}_hot_temp_{zone_number}")
        hot_rows.append({"ZoneNumber": zone_number, "ZoneName": zone_name, "Temperature": temp, "DisplayOrder": str(order)})

    with st.expander("Robot important parameters / 机械手重要参数", expanded=False):
        robot_records: list[dict[str, object]] = []
        if isinstance(current_robot, pd.DataFrame) and not current_robot.empty:
            current_robot = current_robot.copy()
            current_robot["_Sort"] = pd.to_numeric(current_robot["DisplayOrder"], errors="coerce").fillna(999)
            robot_records = current_robot.sort_values(["_Sort", "RobotParameterID"]).head(5).drop(columns=["_Sort"], errors="ignore").to_dict("records")
        while len(robot_records) < 5:
            robot_records.append({})
        robot_rows: list[dict[str, str]] = []
        for index, existing_robot in enumerate(robot_records[:5], start=1):
            cols = st.columns([1, 1])
            with cols[0]:
                name = st.text_input("Parameter name", value=str(existing_robot.get("ParameterName", "") or ""), key=f"{key_prefix}_robot_name_{index}")
            with cols[1]:
                value = st.text_input("Value / content", value=str(existing_robot.get("ParameterValue", "") or ""), key=f"{key_prefix}_robot_value_{index}")
            robot_rows.append({"ParameterName": name, "ParameterValue": value, "DisplayOrder": str(index)})

    values["Notes"] = st.text_area("Parameter notes / 参数备注", value=str(current.get("Notes", "") or ""), key=f"{key_prefix}_setting_notes")
    change_reason = st.text_input("Change reason / 修改原因", key=f"{key_prefix}_change_reason")
    return values, hot_rows, robot_rows, change_reason


def render_mould_notes_section(mould_number: str, row: pd.Series) -> None:
    st.markdown(f"### {t('moulds.notes_edit_label')}")
    current_notes = str(row.get("Notes", "") or "")
    if can_edit_mould_notes():
        with st.form(f"mould_notes_{mould_number}"):
            new_notes = st.text_area(t("moulds.notes_edit_label"), value=current_notes, height=120)
            submitted = st.form_submit_button(t("moulds.save_notes"))
        if submitted:
            try:
                upsert_mould_notes(current_user()["username"], mould_number, new_notes)
                st.success("Notes saved with history.")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))
    else:
        st.text_area(t("moulds.notes_edit_label"), value=current_notes, disabled=True, height=120)


def render_mould_compatibility_section(mould_number: str, all_machine_ids: list[str]) -> None:
    st.markdown(f"### {t('moulds.compatible_machines')}")
    current = get_mould_compatible_machine_ids(mould_number)
    st.write(", ".join(localized_machine_text(machine) for machine in current) if current else t("common.not_set"))
    if not can_edit_mould_compatibility():
        return
    with st.form(f"mould_compat_{mould_number}"):
        selected: list[str] = []
        cols = st.columns(3)
        for index, machine_id in enumerate(all_machine_ids):
            with cols[index % 3]:
                if st.checkbox(f"{t('common.machine')} {machine_id}", value=machine_id in current, key=f"compat_{mould_number}_{machine_id}"):
                    selected.append(machine_id)
        preferred_options = [""] + selected
        preferred = st.selectbox("Preferred machine / 优先机器", preferred_options, format_func=lambda value: "None" if not value else f"Machine {value}")
        compat_note = st.text_area("Compatibility note / 适配备注")
        submitted = st.form_submit_button("Save compatible machines / 保存适配机器")
    if submitted:
        try:
            set_mould_machine_compatibility(
                current_user()["username"],
                mould_number,
                selected,
                preferred_machine=preferred,
                notes_by_machine={machine: compat_note for machine in selected if compat_note},
            )
            st.success("Compatible machines saved.")
            st.rerun()
        except ValueError as exc:
            st.error(str(exc))


def render_mould_machine_parameters_section(mould_number: str, all_machine_ids: list[str]) -> None:
    st.markdown("### Mould + machine parameters / 模具与机器参数")
    compatible = get_mould_compatible_machine_ids(mould_number)
    machine_options = compatible or all_machine_ids
    if not compatible:
        st.info("No compatible machines are set. Set compatibility first; editable users can still choose a machine to create the first parameter set.")
    if not machine_options:
        st.warning("No machines are available.")
        return
    selected_machine = st.selectbox(
        "View parameter configuration / 查看参数配置",
        machine_options,
        key=f"param_machine_{mould_number}",
        format_func=lambda value: f"Machine {value}",
    )
    bundle = get_mould_machine_parameter_bundle(mould_number, selected_machine)
    setting = bundle.get("setting")
    edit_key = f"edit_params_{mould_number}_{selected_machine}"
    if setting and not st.session_state.get(edit_key):
        render_mould_setting_view(bundle, full=True)
        if can_edit_mould_parameters() and st.button("Edit parameters / 编辑参数", key=f"edit_btn_{mould_number}_{selected_machine}"):
            st.session_state[edit_key] = True
            st.rerun()
    elif not setting and not st.session_state.get(edit_key):
        st.warning(f"This mould has no parameter configuration for Machine {selected_machine}.")
        if can_edit_mould_parameters() and st.button("Create parameter configuration / 创建参数配置", key=f"create_param_{mould_number}_{selected_machine}"):
            st.session_state[edit_key] = True
            st.rerun()
    else:
        current_values = dict(setting or {})
        with st.form(f"setting_form_{mould_number}_{selected_machine}"):
            values, hot_rows, robot_rows, change_reason = collect_mould_setting_form(current_values, f"{mould_number}_{selected_machine}", bundle.get("hot_runner"), bundle.get("robot"))
            cols = st.columns(2)
            with cols[0]:
                cancel = st.form_submit_button("Cancel / 取消")
            with cols[1]:
                save = st.form_submit_button("Save parameters / 保存参数", type="primary")
        if cancel:
            st.session_state[edit_key] = False
            st.rerun()
        if save:
            try:
                result = save_mould_machine_setting(
                    current_user()["username"],
                    mould_number,
                    selected_machine,
                    values,
                    hot_runner_rows=hot_rows,
                    robot_rows=robot_rows,
                    change_reason=change_reason,
                )
                warnings = result.get("warnings", [])
                st.success(f"Parameters saved as V{result.get('version')}.")
                for warning in warnings:
                    st.warning(warning)
                st.session_state[edit_key] = False
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))

def production_table_page() -> None:
    show_flash("production_table_flash")
    if not (has_permission("production.view") or can_edit_production()):
        st.error(t("production.permission_denied"))
        return
    st.title(t("production.title"))
    catalog = get_product_catalog()
    if catalog.empty:
        st.warning(t("production.no_catalog"))
        return

    st.caption(t("production.caption"))
    filtered = apply_catalog_search_filters(catalog, "production_catalog")

    st.markdown(t("production.matched_products", count=len(filtered)))
    if filtered.empty:
        st.warning(t("production.no_matching_product"))
        return

    filtered = filtered.reset_index(drop=True)
    choices = [product_choice_label(row, i) for i, row in enumerate(filtered.itertuples())]
    selected_choice = st.selectbox(t("production.select_product"), choices)
    selected_index = choices.index(selected_choice)
    selected = filtered.iloc[selected_index].to_dict()

    material_text = selected.get("MainMaterial", "")
    additive_text = selected.get("Additive", "")
    note_pairs = [
        ("Packaging Type", packaging_type_from_catalog(selected)),
        ("Packaging Option", packaging_option_text(selected)),
        ("Additional Packaging", additional_packaging_display(selected.get("AdditionalPackaging"))),
        ("Carton/Unit/Stack", selected.get("CartonUnitStackQty", "")),
        ("Pallet Qty", selected.get("PalletQty", "")),
        ("Cycle Time", selected.get("CycleTime", "")),
        ("Shot Weight", selected.get("ShotWeight", "")),
        ("Cavity", selected.get("Cavity", "")),
        ("Label", selected.get("HasLabel", "")),
        ("Pallet Bag", selected.get("PalletBag", "")),
        ("Pallet Type", selected.get("PalletType", "")),
        ("Wrap Pallet", selected.get("WrapPallet", "")),
        ("Food Application", selected.get("FoodApplication", "")),
        ("Corner Protector", selected.get("CornerProtector", "")),
        ("Instructions", selected.get("AdditionalInstructions", "")),
    ]
    notes = " | ".join(
        f"{label}: {value}"
        for label, value in note_pairs
        if is_meaningful(value, show_no=(label == "Additional Packaging"))
    )

    detail_col, forecast_col = st.columns([1.25, 1])
    with detail_col:
        show_product_detail_native(selected, notes, material_text, additive_text)
    with forecast_col:
        show_latest_forecast_for_product(
            selected.get("Item", ""),
            selected.get("ProductDetail", "") or selected.get("Item", ""),
        )

    if not can_edit_production():
        st.info(t("production.read_only_info"))
        return

    st.subheader(t("production.save_to_machine"))
    production = get_production()
    all_machine_ids = registered_machine_ids(production)
    primary_mould = active_mould_for_product(selected.get("Item", ""))
    default_mould_number = primary_mould.get("MouldNumber", "")
    compatible_machine_ids = get_mould_compatible_machine_ids(default_mould_number) if default_mould_number else []
    show_all_machines = False
    if compatible_machine_ids:
        st.caption("Compatible machines for the linked mould: " + ", ".join(f"Machine {machine}" for machine in compatible_machine_ids))
        if can_force_incompatible_mould():
            show_all_machines = st.checkbox("Developer override: show all machines / 开发者强制显示全部机器", value=False)
    machine_options = all_machine_ids if show_all_machines or not compatible_machine_ids else compatible_machine_ids
    with st.form("save_catalog_to_machine"):
        machine_id = st.selectbox("Machine / 机器", machine_options)
        status_options = status_options_for_machine(production, machine_id, PRODUCTION_ENTRY_STATUSES)
        show_running_warning_if_needed(production, machine_id)
        machine_rows = production[production["MachineID"].astype(str) == str(machine_id)]
        sequence_default = max([int_value(value) for value in machine_rows["Sequence"].tolist()] + [0]) + 1
        sequence = st.number_input("Sequence / 顺序", min_value=1, step=1, value=sequence_default)
        status = st.selectbox(
            "Status / 状态",
            status_options,
            format_func=production_status_label,
        )
        planned_qty = st.number_input("Planned quantity / 计划数量", min_value=0, step=1)
        completed_qty = st.number_input("Completed quantity / 已完成数量", min_value=0, step=1)
        qc_notes = st.text_input(
            t("production.qc_details"),
            placeholder=t("production.qc_details_placeholder"),
            key="production_qc_notes",
        )
        mould_number = st.text_input("Mould number / 模具编号", value=default_mould_number)
        show_product_mould_hint(selected.get("Item", ""), mould_number)
        override_reason = ""
        if can_force_incompatible_mould() and mould_machine_incompatibility(mould_number, machine_id):
            override_reason = st.text_input("Developer override reason / 开发者强制选择原因")
        extra_notes = st.text_area("Extra production notes / 额外生产备注")
        submitted = st.form_submit_button("Save Production Schedule / 保存生产表")

    render_production_mould_parameter_summary(selected.get("Item", ""), machine_id, mould_number)

    if submitted:
        machine_row = machine_rows.iloc[0].to_dict() if not machine_rows.empty else {"MachineName": str(machine_id)}
        final_notes = notes
        if extra_notes:
            final_notes = f"{final_notes} | Extra: {extra_notes}" if final_notes else extra_notes
        try:
            schedule_id = update_production_record(
                current_user()["username"],
                machine_id,
                {
                    "MachineName": machine_row.get("MachineName", "") or str(machine_id),
                    "Sequence": str(sequence),
                    "Status": status,
                    "ProductType": selected.get("ProductType", ""),
                    "Size": selected.get("Size", ""),
                    "ProductCode": selected.get("Item", ""),
                    "ProductName": selected.get("ProductDetail", "") or selected.get("Item", ""),
                    "PlannedQty": str(planned_qty),
                    "CompletedQty": str(completed_qty),
                    "MouldNumber": mould_number,
                    "Material": material_text,
                    "MaterialLocation": selected.get("MaterialLocation", ""),
                    "ColourMasterbatch": f"{selected.get('Colour', '')} / {additive_text}".strip(" /"),
                    "MaterialWeightG": selected.get("MaterialWeightG", ""),
                    "SecondMaterialWeightG": selected.get("SecondMaterialWeightG", ""),
                    "MasterbatchWeightG": selected.get("MasterbatchWeightG", ""),
                    "UnitWeightG": selected.get("UnitWeightG", ""),
                    "Label": selected.get("HasLabel", ""),
                    "PackagingUnit": selected.get("PackagingUnit", ""),
                    "PackagingType": packaging_type_from_catalog(selected),
                    "PackagingOption": selected.get("PackagingOption", ""),
                    "CartonUnitStackQty": selected.get("CartonUnitStackQty", ""),
                    "PalletQty": selected.get("PalletQty", ""),
                    "AdditionalPackaging": selected.get("AdditionalPackaging", ""),
                    "PalletBag": selected.get("PalletBag", ""),
                    "PalletType": selected.get("PalletType", ""),
                    "WrapPallet": selected.get("WrapPallet", ""),
                    "FoodApplication": selected.get("FoodApplication", ""),
                    "CornerProtector": selected.get("CornerProtector", ""),
                    "InventoryLocationID": selected.get("InventoryLocationID", ""),
                    "AdditionalInstructions": selected.get("AdditionalInstructions", ""),
                    "QCRequired": "YES" if qc_required_for_save(qc_notes) else "NO",
                    "QCNotes": qc_notes_for_save(qc_notes),
                    "Notes": final_notes,
                    "_MouldCompatibilityOverrideReason": override_reason,
                },
            )
            set_flash(
                "production_table_flash",
                f"Production schedule saved: {schedule_id}. Duplicate submissions are blocked automatically.",
            )
            st.rerun()
        except ValueError as exc:
            st.error(str(exc))


def mould_page() -> None:
    st.title("Mould Management / 模具管理")
    moulds = get_moulds()
    links = get_product_mould_links()
    production = get_production()
    all_machine_ids = registered_machine_ids(production)

    st.caption("Search first, then open a mould detail card. Parameters are saved by mould + machine.")
    if can_edit_mould_notes():
        if st.button("Clean ProductCatalog import notes / 清理导入说明", key="clean_imported_mould_notes"):
            changed = clean_imported_mould_notes(current_user()["username"])
            st.success(f"Cleaned {changed} mould note(s).")
            st.rerun()

    search_col, status_col, machine_col = st.columns([2, 1, 1])
    with search_col:
        keyword = st.text_input("Search mould number, mould name, series, product code/name")
    with status_col:
        status_filter = st.selectbox("Status", ["All", "waiting", "running", "maintenance"])
    with machine_col:
        machine_options = ["All"] + all_machine_ids
        machine_filter = st.selectbox("Machine", machine_options)

    if can_edit_moulds():
        with st.expander("Add / edit mould basic info", expanded=False):
            mould_numbers = moulds["MouldNumber"].dropna().astype(str).tolist() if "MouldNumber" in moulds.columns else []
            selected_mould = st.selectbox("Existing mould", ["New mould"] + mould_numbers)
            existing = {}
            if selected_mould != "New mould":
                match = moulds[moulds["MouldNumber"].astype(str) == selected_mould]
                existing = match.iloc[0].to_dict() if not match.empty else {}
            with st.form("mould_form"):
                mould_number = st.text_input("Mould number", value=existing.get("MouldNumber", ""))
                mould_name = st.text_input("Mould name", value=existing.get("MouldName", "") or existing.get("AssociatedProduct", ""))
                mould_type = st.text_input("Mould type", value=existing.get("MouldType", ""))
                mould_size = st.text_input("Size", value=existing.get("MouldSize", ""))
                mould_series = st.text_input("Series", value=existing.get("MouldSeries", ""))
                storage = st.text_input("Storage location", value=existing.get("StorageLocation", ""))
                active = st.checkbox("Active", value=truthy(existing.get("Active", "TRUE")))
                submitted = st.form_submit_button("Save Mould")
            if submitted:
                if not mould_number:
                    st.error("Mould number is required.")
                else:
                    upsert_mould(
                        current_user()["username"],
                        mould_number,
                        {
                            "MouldName": mould_name,
                            "MouldType": mould_type,
                            "MouldSize": mould_size,
                            "MouldSeries": mould_series,
                            "MouldFamily": mould_series or mould_size,
                            "AssociatedProduct": mould_name,
                            "StorageLocation": storage,
                            "Status": existing.get("Status", "Available") or "Available",
                            "ManualStatus": existing.get("ManualStatus", "waiting") or "waiting",
                            "Active": "TRUE" if active else "FALSE",
                            "Notes": existing.get("Notes", ""),
                            "MaintenanceNotes": existing.get("MaintenanceNotes", ""),
                            "CreatedAt": existing.get("CreatedAt", ""),
                            "CreatedBy": existing.get("CreatedBy", current_user()["username"]),
                        },
                    )
                    st.success("Mould record saved.")
                    st.rerun()

    if not keyword and status_filter == "All" and machine_filter == "All":
        st.info("Enter a mould number/name, product code/name, or choose a filter to show results.")
        return

    filtered = moulds.copy()
    if keyword:
        mask = pd.Series(False, index=filtered.index)
        search_columns = [
            "MouldNumber",
            "MouldName",
            "MouldFamily",
            "MouldSize",
            "MouldSeries",
            "MouldType",
            "AssociatedProduct",
            "StorageLocation",
            "Status",
            "IssueDescription",
            "Notes",
            "MaintenanceNotes",
        ]
        for column in search_columns:
            if column in filtered.columns:
                mask = mask | filtered[column].astype(str).str.contains(keyword, case=False, na=False, regex=False)
        if not links.empty:
            linked_codes = links[
                links["ProductCode"].astype(str).str.contains(keyword, case=False, na=False, regex=False)
                | links["ProductName"].astype(str).str.contains(keyword, case=False, na=False, regex=False)
            ]["MouldNumber"].astype(str)
            mask = mask | filtered["MouldNumber"].astype(str).isin(set(linked_codes))
        filtered = filtered[mask]

    cards = []
    for _, row in filtered.iterrows():
        mould_number = str(row.get("MouldNumber", "") or "")
        status = resolve_mould_status(mould_number)
        if status_filter != "All" and status.get("status") != status_filter:
            continue
        compatible = get_mould_compatible_machine_ids(mould_number)
        if machine_filter != "All" and status.get("machine_id") != machine_filter and machine_filter not in compatible:
            continue
        linked = get_mould_linked_products(mould_number)
        cards.append((row, status, linked, compatible))

    if not cards:
        st.info("No moulds match the current search/filter.")
        return
    if len(cards) > 20:
        st.warning(f"Showing first 20 of {len(cards)} results. Add more search text to narrow the list.")
        cards = cards[:20]

    for row, status, linked, compatible in cards:
        mould_number = str(row.get("MouldNumber", "") or "")
        mould_name = str(row.get("MouldName", "") or row.get("AssociatedProduct", "") or "-")
        compatible_text = ", ".join(f"Machine {machine}" for machine in compatible) if compatible else "Not set"
        title = f"{mould_number} | {mould_name} | {status.get('status_label')} | {status.get('location')} | Linked products: {len(linked)}"
        with st.expander(title, expanded=len(cards) == 1):
            st.markdown(
                f"""
                <div class="info-card">
                    {card_field("Mould Number", mould_number)}
                    {card_field("Mould Name", mould_name)}
                    {card_field("Type", row.get("MouldType"))}
                    {card_field("Size", row.get("MouldSize"))}
                    {card_field("Series", row.get("MouldSeries"))}
                    {card_field("Status", status.get("status_label"))}
                    {card_field("Location", status.get("location"))}
                    {card_field("Compatible machines", compatible_text)}
                    {card_field("Last Updated", row.get("LastUpdated"))}
                </div>
                """,
                unsafe_allow_html=True,
            )

            render_mould_compatibility_section(mould_number, all_machine_ids)
            render_mould_machine_parameters_section(mould_number, all_machine_ids)
            render_mould_notes_section(mould_number, row)

            st.markdown("**Linked products / 挂钩产品**")
            if linked.empty:
                st.info("No linked products yet.")
            else:
                display_cols = [col for col in [
                    "ProductCode", "ProductName", "ProductType", "Colour", "HasLabel",
                    "IsPrimary", "MachineID", "Status"
                ] if col in linked.columns]
                st.dataframe(linked[display_cols], use_container_width=True, hide_index=True)

            st.markdown("**Maintenance history / 维修历史**")
            maintenance = get_mould_maintenance_history()
            mould_maintenance = maintenance[
                maintenance["MouldNumber"].astype(str).str.casefold().eq(mould_number.casefold())
                & ~maintenance["IsDeleted"].astype(str).str.casefold().isin({"true", "yes", "1"})
            ].copy()
            if mould_maintenance.empty:
                st.info("No maintenance records.")
            else:
                st.dataframe(mould_maintenance.sort_values("CreatedAt", ascending=False), use_container_width=True, hide_index=True)

            if can_add_mould_maintenance():
                with st.form(f"add_maintenance_{mould_number}"):
                    st.markdown("**Add maintenance record**")
                    technician = st.text_input("Technician name", key=f"tech_{mould_number}")
                    maintenance_content = st.text_area("Maintenance content", key=f"content_{mould_number}")
                    set_maintenance = st.checkbox("Set mould status to Maintenance", value=True, key=f"set_maint_{mould_number}")
                    add_submitted = st.form_submit_button("Save maintenance record")
                if add_submitted:
                    try:
                        add_mould_maintenance_record(
                            current_user()["username"],
                            mould_number,
                            technician,
                            maintenance_content,
                            set_maintenance=set_maintenance,
                        )
                        st.success("Maintenance record saved and locked.")
                        st.rerun()
                    except ValueError as exc:
                        st.error(str(exc))

            open_records = mould_maintenance[
                ~mould_maintenance["Status"].astype(str).str.casefold().isin({"completed", "closed", "cancelled"})
            ]
            if can_complete_mould_maintenance() and not open_records.empty:
                with st.form(f"complete_maintenance_{mould_number}"):
                    selected_id = st.selectbox("Open maintenance record", open_records["MaintenanceID"].astype(str).tolist())
                    completion_note = st.text_area("Completion note")
                    complete_submitted = st.form_submit_button("Mark maintenance completed")
                if complete_submitted:
                    try:
                        complete_mould_maintenance_record(current_user()["username"], selected_id, completion_note)
                        st.success("Maintenance completed.")
                        st.rerun()
                    except ValueError as exc:
                        st.error(str(exc))

            st.markdown("**Parameter history / 参数修改历史**")
            history = get_mould_machine_settings_history()
            mould_history = history[history["MouldNumber"].astype(str).str.casefold().eq(mould_number.casefold())].copy()
            if mould_history.empty:
                st.info("No parameter history.")
            else:
                show_cols = [col for col in ["ChangedAt", "ChangedBy", "MachineID", "Version", "ChangeReason"] if col in mould_history.columns]
                st.dataframe(mould_history.sort_values("ChangedAt", ascending=False)[show_cols], use_container_width=True, hide_index=True)

            st.markdown("**Notes history / 备注修改历史**")
            notes_history = get_mould_notes_history()
            mould_notes_history = notes_history[notes_history["MouldNumber"].astype(str).str.casefold().eq(mould_number.casefold())].copy()
            if mould_notes_history.empty:
                st.info("No notes history.")
            else:
                st.dataframe(mould_notes_history.sort_values("ChangedAt", ascending=False), use_container_width=True, hide_index=True)

            st.markdown("**Maintenance audit log / 维修审计日志**")
            audit = get_mould_maintenance_audit_log()
            mould_audit = audit[audit["MouldNumber"].astype(str).str.casefold().eq(mould_number.casefold())].copy()
            if mould_audit.empty:
                st.info("No maintenance audit log.")
            else:
                st.dataframe(mould_audit.sort_values("Timestamp", ascending=False), use_container_width=True, hide_index=True)

    st.subheader("Setup diagrams / 安装图")
    st.info("Water pipe, air pipe, and other setup diagrams will be completed in a future version.")



def safe_excel_sheet_name(title: object, fallback: str = "Sheet") -> str:
    name = re.sub(r"[\\/*?:\[\]]", " ", str(title or "").strip())
    name = re.sub(r"\s+", " ", name).strip(" '")
    return (name or fallback)[:31]


def unique_excel_sheet_name(title: object, used: set[str]) -> str:
    base = safe_excel_sheet_name(title)
    name = base
    counter = 2
    while name in used:
        suffix = f" {counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(name)
    return name

def data_quality_panel() -> None:
    catalog = get_product_catalog()
    moulds = get_moulds()
    links = get_product_mould_links()
    production = get_production()
    media = get_mould_media()
    active_links = links[links["Active"].str.lower().isin(["true", "yes", "1"])]
    linked_codes = set(active_links["ProductCode"])
    unlinked = catalog[~catalog["Item"].isin(linked_codes) & catalog["Item"].ne("")]
    pallet = pd.to_numeric(catalog["PalletQty"], errors="coerce")
    bad_pallet = catalog[pallet.isna() | pallet.le(0)]
    mould_numbers = set(moulds["MouldNumber"])
    missing_mould = active_links[~active_links["MouldNumber"].isin(mould_numbers)]
    invalid_running = production[
        production["Status"].isin(["Running", "Next"])
        & ~production["ProductCode"].isin(linked_codes)
    ]
    reports = {
        "Products without active mould link": unlinked,
        "Products with missing/invalid PalletQty": bad_pallet,
        "Links to missing mould": missing_mould,
        "Running/Next without active link": invalid_running,
        "Moulds missing location": moulds[moulds["StorageLocation"].eq("")],
        "Moulds missing status": moulds[moulds["Status"].eq("")],
    }
    for title, frame in reports.items():
        with st.expander(f"{title}: {len(frame)}"):
            st.dataframe(frame, use_container_width=True, hide_index=True)
    missing_local = []
    media_root = os.getenv("MOULD_MEDIA_LOCAL_ROOT", str(os.path.join(os.path.dirname(__file__), "data", "mould_media")))
    for _, row in media.iterrows():
        relative = str(row.get("LocalRelativePath", "")).strip()
        if relative and not os.path.exists(os.path.join(media_root, relative)):
            missing_local.append(row.to_dict())
    st.metric("Media records with missing local file", len(missing_local))
    output = BytesIO()
    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for title, frame in reports.items():
            frame.to_excel(writer, sheet_name=unique_excel_sheet_name(title, used_sheet_names), index=False)
    st.download_button("Export Data Quality Report", output.getvalue(), "factory_data_quality.xlsx")


def qr_poster_png_bytes(title: str, subtitle: str, url: str, error_correction: str = "Q") -> bytes:
    image = qr_poster_image(title, subtitle, url, error_correction)
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


def build_qr_zip_bytes(base_url: str, error_correction: str = "Q") -> bytes:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for title, subtitle, url, filename in build_qr_specs(base_url.rstrip("/")):
            archive.writestr(filename, qr_poster_png_bytes(title, subtitle, url, error_correction))
    return buffer.getvalue()


def _legacy_qr_generator_panel(machine_ids: list[str]) -> None:
    st.subheader("Factory MIS QR Generator")
    try:
        configured_cloud_url = load_settings().mobile_base_url
    except Exception:
        configured_cloud_url = ""
    st.caption("Use Admin URLs for office/back-office access. Use Cloud Mobile URL for production machine and Stock-In QR codes.")
    local_admin_url = st.text_input(
        "Local Admin URL",
        value="http://localhost:8501",
        help="Full local Admin backend on this computer. This is not a mobile production QR URL.",
    ).strip().rstrip("/")
    wifi_admin_url = st.text_input(
        "WiFi Admin URL",
        value="",
        placeholder="http://192.168.x.x:8501",
        help="Full Admin backend over LAN/WiFi. Use for office/admin access only.",
    ).strip().rstrip("/")
    cloud_url = st.text_input(
        "Production Mobile Cloud URL",
        value=configured_cloud_url,
        placeholder="https://factory-mobile.example.com",
        help="Use the fixed Streamlit Cloud deployment URL. Do not use 127.0.0.1 or temporary trycloudflare.com URLs for production QR codes.",
    ).strip().rstrip("/")

    with st.expander("Admin URLs / Local WiFi testing only"):
        st.caption("These open the full Admin backend. They are not production machine/Stock-In QR URLs.")
        if local_admin_url:
            st.code(local_admin_url)
        if wifi_admin_url:
            st.code(wifi_admin_url)
        else:
            st.info("Enter the office PC WiFi/LAN IP if you want an Admin WiFi URL.")

    if not cloud_url:
        st.info("Enter the deployed Mobile Cloud URL before generating production QR codes.")
        return

    qr_type = st.radio(
        "Production Cloud QR type",
        ["Machine Status Overview", "Machine", "Stock-In Request", "All Production QR Codes ZIP"],
        horizontal=False,
    )

    if qr_type == "Machine Status Overview":
        lang = st.radio("Language / 语言", ["en", "zh-CN"], horizontal=True, key="qr_lang_overview")
        target_url = f"{cloud_url}/?page=machine_status&lang={lang}"
        image_bytes = qr_png_bytes(target_url)
        st.code(target_url)
        st.image(image_bytes, caption="Machine Status", width=220)
        st.download_button("Download Machine Status QR", image_bytes, file_name="machine_status.png", mime="image/png")
    elif qr_type == "Machine":
        machine_id = st.selectbox("Machine", machine_ids)
        lang = st.radio("Language / 语言", ["en", "zh-CN"], horizontal=True, key="qr_lang_machine")
        target_url = f"{cloud_url}/?page=machine_status&machine_id={machine_id}&lang={lang}"
        image_bytes = qr_png_bytes(target_url)
        st.code(target_url)
        st.image(image_bytes, caption=f"Machine {machine_id}", width=220)
        st.download_button("Download Machine QR PNG", image_bytes, file_name=f"machine_{machine_id}.png", mime="image/png")
    elif qr_type == "Stock-In Request":
        lang = st.radio("Language / 语言", ["en", "zh-CN"], horizontal=True, key="qr_lang_stock")
        target_url = f"{cloud_url}/?page=stock_in&lang={lang}"
        image_bytes = qr_png_bytes(target_url)
        st.code(target_url)
        st.image(image_bytes, caption="Stock-In Request", width=220)
        st.download_button("Download Stock-In QR PNG", image_bytes, file_name="stock_in_request.png", mime="image/png")
    else:
        zip_bytes = build_qr_zip(cloud_url, machine_ids)
        st.download_button(
            "Download All Production QR Codes ZIP",
            zip_bytes,
            file_name="factory_mobile_cloud_qr_codes.zip",
            mime="application/zip",
        )

    with st.expander("Optional WiFi mobile test URLs"):
        st.caption("Local WiFi testing only. Production QR codes should use the fixed Cloud Mobile URL above.")
        if wifi_admin_url:
            wifi_mobile_base = wifi_admin_url.replace(":8501", ":8502")
            st.code(f"{wifi_mobile_base}/?page=machine_status&lang=zh-CN")
            st.code(f"{wifi_mobile_base}/?page=stock_in&lang=zh-CN")
        else:
            st.info("Enter WiFi Admin URL first if you need local WiFi mobile test URLs.")


def qr_generator_panel(machine_ids: list[str]) -> None:
    st.subheader("Factory MIS QR Generator")
    try:
        legacy_cloud_url = load_settings().mobile_base_url
    except Exception:
        legacy_cloud_url = ""
    configured_cloud_url = read_qr_env_base_url() or legacy_cloud_url

    st.caption(
        "Generate production mobile QR posters from the local Admin page. "
        "Use the fixed Streamlit Cloud mobile URL for machine and Stock-In QR codes."
    )
    local_admin_url = st.text_input(
        "Local Admin URL",
        value="http://localhost:8501",
        help="Full local Admin backend on this computer. This is not a mobile production QR URL.",
    ).strip().rstrip("/")
    wifi_admin_url = st.text_input(
        "WiFi Admin URL",
        value="",
        placeholder="http://192.168.x.x:8501",
        help="Full Admin backend over LAN/WiFi. Use for office/admin access only.",
    ).strip().rstrip("/")
    cloud_url = st.text_input(
        "Production Mobile Cloud URL",
        value=configured_cloud_url,
        placeholder="https://your-app.streamlit.app",
        help="Use the fixed Streamlit Cloud deployment URL. You can also save it as MOBILE_PUBLIC_BASE_URL in .env.",
    ).strip().rstrip("/")
    error_correction = st.radio(
        "QR error correction",
        ["Q", "M"],
        index=0,
        horizontal=True,
        help="Q is more robust for printed posters; M creates a slightly denser code.",
    )

    with st.expander("Admin URLs / Local WiFi testing only"):
        st.caption("These open the full Admin backend. They are not production machine/Stock-In QR URLs.")
        if local_admin_url:
            st.code(local_admin_url)
        if wifi_admin_url:
            st.code(wifi_admin_url)
        else:
            st.info("Enter the office PC WiFi/LAN IP if you want an Admin WiFi URL.")

    if not cloud_url:
        st.info("Enter the deployed Mobile Cloud URL before generating production QR codes.")
        return

    qr_type = st.radio(
        "Production Cloud QR type",
        ["Machine Status Overview", "Machine", "Stock-In Request", "All Production QR Codes ZIP"],
        horizontal=False,
    )

    if qr_type == "Machine Status Overview":
        lang = st.radio("Language", ["en", "zh-CN"], horizontal=True, key="qr_lang_overview")
        lang_suffix = "zh" if lang == "zh-CN" else "en"
        target_url = make_qr_url(cloud_url, "machine_status", lang)
        image_bytes = qr_poster_png_bytes("Machine Status", "Scan to view all machines", target_url, error_correction)
        st.code(target_url)
        st.image(image_bytes, caption="Machine Status", width=260)
        st.download_button(
            "Download Machine Status QR",
            image_bytes,
            file_name=f"machine_status_{lang_suffix}_qr.png",
            mime="image/png",
        )
    elif qr_type == "Machine":
        if not machine_ids:
            st.warning("No machines found. Add or publish machine data before generating machine QR codes.")
            return
        machine_id = st.selectbox("Machine", machine_ids)
        lang = st.radio("Language", ["en", "zh-CN"], horizontal=True, key="qr_lang_machine")
        lang_suffix = "zh" if lang == "zh-CN" else "en"
        target_url = make_qr_url(cloud_url, "machine_status", lang, machine_id=machine_id)
        image_bytes = qr_poster_png_bytes(f"Machine {machine_id}", "Scan to view this machine", target_url, error_correction)
        st.code(target_url)
        st.image(image_bytes, caption=f"Machine {machine_id}", width=260)
        st.download_button(
            "Download Machine QR PNG",
            image_bytes,
            file_name=f"machine_{safe_qr_filename_part(machine_id)}_{lang_suffix}_qr.png",
            mime="image/png",
        )
    elif qr_type == "Stock-In Request":
        lang = st.radio("Language", ["en", "zh-CN"], horizontal=True, key="qr_lang_stock")
        lang_suffix = "zh" if lang == "zh-CN" else "en"
        target_url = make_qr_url(cloud_url, "stock_in", lang)
        image_bytes = qr_poster_png_bytes("Stock-In Request", "Scan to submit stock-in request", target_url, error_correction)
        st.code(target_url)
        st.image(image_bytes, caption="Stock-In Request", width=260)
        st.download_button(
            "Download Stock-In QR PNG",
            image_bytes,
            file_name=f"stock_in_{lang_suffix}_qr.png",
            mime="image/png",
        )
    else:
        zip_bytes = build_qr_zip_bytes(cloud_url, error_correction)
        if st.button("Generate PNG files and ZIP in public_qr / Generate local files"):
            QR_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            png_paths = []
            for title, subtitle, url, filename in build_qr_specs(cloud_url):
                path = QR_OUTPUT_DIR / filename
                save_qr_poster(title, subtitle, url, path, error_correction)
                png_paths.append(path)
            zip_path = create_qr_zip(QR_OUTPUT_DIR, png_paths)
            st.success(f"Generated {len(png_paths)} PNG files and ZIP: {zip_path}")
        st.download_button(
            "Download All Production QR Codes ZIP",
            zip_bytes,
            file_name="factory_mobile_cloud_qr_codes.zip",
            mime="application/zip",
        )
        st.caption(f"Default local output folder: {QR_OUTPUT_DIR}")

    with st.expander("Optional WiFi mobile test URLs"):
        st.caption("Local WiFi testing only. Production QR codes should use the fixed Cloud Mobile URL above.")
        if wifi_admin_url:
            wifi_mobile_base = wifi_admin_url.replace(":8501", ":8502")
            st.code(f"{wifi_mobile_base}/?page=machine_status&lang=zh-CN")
            st.code(f"{wifi_mobile_base}/?page=stock_in&lang=zh-CN")
        else:
            st.info("Enter WiFi Admin URL first if you need local WiFi mobile test URLs.")


PERMISSION_LABELS = {
    "CanEditProduction": "Can modify machine / production status",
    "CanEditInventory": "Can modify inventory directly",
    "CanEditMoulds": "Can create/edit mould records",
    "CanManageUsers": "Can manage user accounts",
    "CanLinkProductMould": "Can manage product-mould links",
    "CanManageMouldIssues": "Can manage mould issues",
    "CanUploadMouldMedia": "Can upload mould media",
    "CanManageMouldMedia": "Can manage mould media",
    "CanStockIn": "Can perform Stock-In",
    "CanManageLooseGoods": "Can manage Loose Goods",
    "CanAddMouldMaintenance": "Can add mould maintenance records",
    "CanCompleteMouldMaintenance": "Can mark mould maintenance completed",
    "CanEditLockedMouldMaintenance": "Developer only: edit locked maintenance records",
}


def truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"true", "yes", "y", "1"}


def user_management_panel() -> None:
    st.subheader("User Accounts / 权限管理")
    users = get_users()
    visible_columns = ["Username", "Role", "Active", *PERMISSION_COLUMNS]
    st.dataframe(users[visible_columns], use_container_width=True, hide_index=True)

    editable_users = users[users["Role"] != "Developer"].copy()
    if not editable_users.empty:
        st.markdown("**Edit existing account / 修改现有账号**")
        selected_user = st.selectbox("Account", editable_users["Username"].tolist())
        row = editable_users[editable_users["Username"] == selected_user].iloc[0]
        with st.form("edit_user_permissions"):
            active = st.checkbox("Active", value=truthy(row.get("Active")))
            role_options = ["Administrator", "Technical Manager", "Viewer"]
            current_role = row.get("Role", "Administrator")
            role = st.selectbox("Role", role_options, index=role_options.index(current_role) if current_role in role_options else 0)
            defaults = role_default_permissions(role)
            permissions = {
                column: st.checkbox(
                    label,
                    value=truthy(row.get(column)) if str(row.get(column, "")).strip() else defaults.get(column, False),
                    key=f"edit_{column}",
                )
                for column, label in PERMISSION_LABELS.items()
            }
            submitted = st.form_submit_button("Save Account Permissions")
        if submitted:
            try:
                upsert_user(current_user()["username"], selected_user, "", role, active, permissions)
                st.success("Account permissions saved.")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))

    st.markdown("**Create account / 创建账号**")
    with st.form("create_admin_user"):
        new_username = st.text_input("New username")
        new_role = st.selectbox("Role", ["Administrator", "Technical Manager", "Viewer"])
        new_password = st.text_input("Password", type="password")
        confirm_password = st.text_input("Confirm password", type="password")
        defaults = role_default_permissions(new_role)
        new_permissions = {
            column: st.checkbox(label, value=defaults.get(column, False), key=f"new_{column}")
            for column, label in PERMISSION_LABELS.items()
        }
        create_submitted = st.form_submit_button("Create Account")
    if create_submitted:
        if new_password != confirm_password:
            st.error("Passwords do not match.")
        else:
            try:
                upsert_user(current_user()["username"], new_username, new_password, new_role, True, new_permissions)
                st.success("Account created.")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))


def production_change_review_panel() -> None:
    st.subheader("Production Change Review / 生产变更审核")
    st.warning("Mobile reports are not official production-plan changes until an admin confirms and applies them.")
    changes = get_production_change_requests()
    if changes.empty:
        st.info("No production change requests yet.")
        return

    statuses = ["All", "Pending Review", "Admin Confirmed", "Rejected", "Applied to Production Plan"]
    status_filter = st.selectbox("Status filter", statuses, key="production_change_status_filter")
    filtered = changes.copy()
    if status_filter != "All":
        filtered = filtered[filtered["Status"].astype(str).eq(status_filter)]
    search = st.text_input("Search machine, product, reason, reporter", key="production_change_search")
    if search:
        mask = pd.Series(False, index=filtered.index)
        for column in ["MachineNo", "OldProductCode", "OldProductName", "NewProductCode", "NewProductName", "Reason", "ReportedBy", "Status"]:
            mask = mask | filtered[column].astype(str).str.contains(search, case=False, na=False)
        filtered = filtered[mask]

    if filtered.empty:
        st.info("No production change requests match the current filter.")
        return

    summary_cols = [
        "ChangeID",
        "CreatedAt",
        "MachineNo",
        "OldProductName",
        "NewProductName",
        "ChangeTime",
        "ReportedBy",
        "Status",
        "AppliedToPlan",
    ]
    st.dataframe(filtered[summary_cols].sort_values("CreatedAt", ascending=False), use_container_width=True, hide_index=True)

    production = get_production()
    for _, request in filtered.sort_values("CreatedAt", ascending=False).iterrows():
        change_id = str(request.get("ChangeID", ""))
        title = (
            f"{change_id} | Machine {request.get('MachineNo', '-')} | "
            f"{request.get('OldProductName', request.get('OldProductCode', '-'))} -> "
            f"{request.get('NewProductName', request.get('NewProductCode', '-'))} | {request.get('Status', '-')}"
        )
        with st.expander(title, expanded=str(request.get("Status", "")) == "Pending Review"):
            left, right = st.columns(2)
            with left:
                st.markdown("**Reported change**")
                st.write(f"Machine: **{request.get('MachineNo', '-')}**")
                st.write(f"Old: `{request.get('OldProductCode', '')}` {request.get('OldProductName', '')} / {request.get('OldColor', '')}")
                st.write(f"New: `{request.get('NewProductCode', '')}` {request.get('NewProductName', '')} / {request.get('NewColor', '')}")
                st.write(f"Change time: {request.get('ChangeTime', '-')}")
                st.write(f"Reported completed: {request.get('ReportedCompletedQty', '0')}")
                st.write(f"Reported remaining: {request.get('ReportedRemainingQty', '0')}")
            with right:
                st.markdown("**Reason and review**")
                st.write(f"Reason: {request.get('Reason', '-')}")
                st.write(f"Reported by: {request.get('ReportedBy', '-')}")
                st.write(f"Note: {request.get('Note', '-')}")
                st.write(f"Photo: {request.get('PhotoPath', '-') or '-'}")
                st.write(f"Admin note: {request.get('AdminNote', '-') or '-'}")

            machine_no = str(request.get("MachineNo", "")).strip()
            running = production[
                production["MachineID"].astype(str).str.strip().str.casefold().eq(machine_no.casefold())
                & production["Status"].astype(str).str.strip().str.casefold().eq("running")
            ]
            current_row = running.iloc[0].to_dict() if not running.empty else None
            preview = production_change_apply_preview(request.to_dict(), current_row)
            st.markdown("**Apply preview / 应用预览**")
            preview_rows = []
            for field in ["ProductCode", "ProductName", "ColourMasterbatch", "PlannedQty", "CompletedQty", "Status"]:
                preview_rows.append({"Field": field, "Current official value": preview["old"].get(field, ""), "After apply": preview["new"].get(field, "")})
            st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
            if current_row is None:
                st.warning("No current Running production record was found for this machine. Review machine state before applying.")

            status = str(request.get("Status", ""))
            applied = str(request.get("AppliedToPlan", "")).strip().casefold() in {"true", "yes", "1"}
            if not can_edit_production():
                st.info("Your account does not have permission to review or apply production changes.")
                continue

            if status == "Pending Review":
                col_confirm, col_reject = st.columns(2)
                with col_confirm:
                    if st.button("Confirm / 确认", key=f"confirm_change_{change_id}"):
                        try:
                            review_production_change_request(current_user()["username"], change_id, "Admin Confirmed", "")
                            st.success("Request confirmed.")
                            st.rerun()
                        except ValueError as exc:
                            st.error(str(exc))
                with col_reject:
                    with st.form(f"reject_change_form_{change_id}"):
                        admin_note = st.text_area("Reject note / 拒绝原因", key=f"reject_note_{change_id}")
                        reject = st.form_submit_button("Reject / 拒绝")
                    if reject:
                        try:
                            review_production_change_request(current_user()["username"], change_id, "Rejected", admin_note)
                            st.success("Request rejected.")
                            st.rerun()
                        except ValueError as exc:
                            st.error(str(exc))
            elif status == "Admin Confirmed" and not applied:
                confirm_apply = st.checkbox(f"I reviewed the preview and want to apply {change_id}", key=f"apply_confirm_{change_id}")
                if st.button("Apply to Production Plan / 应用到正式生产计划", key=f"apply_change_{change_id}", disabled=not confirm_apply):
                    try:
                        schedule_id = apply_production_change_request(current_user()["username"], change_id)
                        st.success(f"Applied to production schedule {schedule_id}.")
                        st.rerun()
                    except ValueError as exc:
                        st.error(str(exc))
            elif status == "Rejected":
                st.info("Rejected requests cannot be applied.")
            elif status == "Applied to Production Plan" or applied:
                st.success(f"Already applied to schedule {request.get('AppliedScheduleID', '-')}.")



def _performance_display_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    columns = [
        "ProductionPlanID",
        "Status",
        "MachineID",
        "ProductCode",
        "ProductName",
        "PlannedQty",
        "ActualStockInQty",
        "StockInRecordCount",
        "AdjustedProductionHours",
        "MachineRuntimeHours",
        "NonRuntimeHours",
        "ScrapRatePercent",
        "CompletedAt",
    ]
    available = [column for column in columns if column in df.columns]
    return df[available].copy()


def _performance_record_label(row: pd.Series) -> str:
    return (
        f"{row.get('ProductionPlanID', '')} | Machine {row.get('MachineID', '')} | "
        f"{row.get('ProductName', '') or row.get('ProductCode', '')}"
    )


def _performance_editor(record: dict[str, object], user_name: str, form_key: str, allow_archive: bool = True) -> None:
    plan_id = str(record.get("ProductionPlanID", "") or "")
    stock_summary = refresh_production_performance_stock_summary(plan_id)
    calc_weekend = stock_summary.get("NaturalElapsedHours", record.get("NaturalElapsedHours", 0))
    st.markdown(f"### {t('performance.plan_id')}: `{plan_id}`")
    summary_cols = st.columns(4)
    summary_cols[0].metric("Machine", record.get("MachineID", ""))
    summary_cols[1].metric("Planned", record.get("PlannedQty", ""))
    summary_cols[2].metric("Stock-in qty", stock_summary.get("ActualStockInQty", 0))
    summary_cols[3].metric("Records", stock_summary.get("StockInRecordCount", 0))
    if int(stock_summary.get("StockInRecordCount", 0) or 0) == 1:
        st.warning(t("performance.warning_single_stock"))

    with st.form(form_key):
        st.caption(f"{record.get('ProductCode', '')} - {record.get('ProductName', '')}")
        st.markdown(f"#### {t('performance.time_section')}")
        time_cols = st.columns(3)
        weekend_options = ["No", "Yes", "Partial"]
        weekend_current = str(record.get("WeekendProductionMode", "No") or "No")
        weekend_mode = time_cols[0].selectbox(
            t("performance.weekend_mode"),
            weekend_options,
            index=weekend_options.index(weekend_current) if weekend_current in weekend_options else 0,
        )
        weekend_included = time_cols[1].number_input(
            t("performance.weekend_included"),
            value=float(record.get("WeekendHoursIncluded") or 0),
            step=0.25,
        )
        correction = time_cols[2].number_input(
            t("performance.time_correction"),
            value=float(record.get("TimeCorrectionHours") or 0),
            step=0.25,
        )
        correction_reason = st.text_input(t("performance.time_reason"), value=str(record.get("TimeCorrectionReason", "") or ""))
        runtime = st.number_input(
            t("performance.runtime"),
            value=float(record.get("MachineRuntimeHours") or 0),
            min_value=0.0,
            step=0.25,
        )

        st.markdown(f"#### {t('performance.material_section')}")
        material_cols = st.columns(4)
        good_qty = material_cols[0].number_input(
            t("performance.good_qty"),
            value=float(record.get("GoodQty") or stock_summary.get("ActualStockInQty") or 0),
            min_value=0.0,
            step=1.0,
        )
        scrap_qty = material_cols[1].number_input(
            t("performance.scrap_qty"),
            value=float(record.get("ScrapQty") or 0),
            min_value=0.0,
            step=1.0,
        )
        unit_weight = material_cols[2].number_input(
            "Standard unit weight g",
            value=float(record.get("StandardUnitWeightG") or 0),
            min_value=0.0,
            step=0.1,
        )
        actual_material = material_cols[3].number_input(
            t("performance.actual_material"),
            value=float(record.get("ActualMaterialKg") or 0),
            min_value=0.0,
            step=0.1,
        )
        operator = st.text_input("Operator", value=str(record.get("Operator", "") or ""))
        shift = st.text_input("Shift", value=str(record.get("Shift", "") or ""))
        exception_notes = st.text_area(t("performance.exception_notes"), value=str(record.get("ExceptionNotes", "") or ""), height=100)
        notes = st.text_area(t("production.notes"), value=str(record.get("Notes", "") or ""), height=80)
        no_stock_confirm = True
        if int(stock_summary.get("StockInRecordCount", 0) or 0) == 0:
            no_stock_confirm = st.checkbox(t("performance.no_stock_confirm"), value=False)

        values = {
            "WeekendProductionMode": weekend_mode,
            "WeekendHoursIncluded": weekend_included,
            "TimeCorrectionHours": correction,
            "TimeCorrectionReason": correction_reason,
            "MachineRuntimeHours": runtime,
            "GoodQty": good_qty,
            "ScrapQty": scrap_qty,
            "StandardUnitWeightG": unit_weight,
            "ActualMaterialKg": actual_material,
            "Operator": operator,
            "Shift": shift,
            "ExceptionNotes": exception_notes,
            "Notes": notes,
            "ActualStockInQty": stock_summary.get("ActualStockInQty", 0),
            "StockInRecordCount": stock_summary.get("StockInRecordCount", 0),
            "FirstStockInTime": stock_summary.get("FirstStockInTime", ""),
            "LastStockInTime": stock_summary.get("LastStockInTime", ""),
            "NaturalElapsedHours": stock_summary.get("NaturalElapsedHours", calc_weekend),
            "CalculatedWeekendHours": record.get("CalculatedWeekendHours", 0),
        }
        metrics = calculate_production_performance_metrics(values)
        metric_cols = st.columns(4)
        metric_cols[0].metric("Adjusted hours", metrics.get("AdjustedProductionHours", 0))
        metric_cols[1].metric(t("performance.non_runtime"), metrics.get("NonRuntimeHours", ""))
        metric_cols[2].metric("Theoretical kg", metrics.get("TheoreticalMaterialKg", ""))
        metric_cols[3].metric("Scrap %", metrics.get("ScrapRatePercent", ""))
        if isinstance(metrics.get("NonRuntimeHours"), (int, float)) and float(metrics["NonRuntimeHours"]) < 0:
            st.warning(t("performance.warning_negative"))

        save_col, archive_col = st.columns(2)
        save_clicked = save_col.form_submit_button(t("performance.save_draft"), use_container_width=True)
        archive_clicked = archive_col.form_submit_button(t("performance.archive"), type="primary", use_container_width=True, disabled=not allow_archive)
        if save_clicked or archive_clicked:
            if archive_clicked and int(stock_summary.get("StockInRecordCount", 0) or 0) == 0 and not no_stock_confirm:
                st.error(t("performance.no_stock_confirm"))
                return
            try:
                update_production_performance_record(user_name, plan_id, values, archive=archive_clicked)
            except Exception as exc:
                st.error(str(exc))
            else:
                st.success(t("performance.archived_saved") if archive_clicked else t("performance.saved"))
                st.rerun()


def production_performance_page() -> None:
    st.title(t("performance.title"))
    user = current_user()
    user_name = str(user.get("username", "") if user else "")
    if not is_developer():
        st.error(t("performance.permission_denied"))
        return

    if st.button(t("performance.backfill")):
        try:
            count = backfill_production_performance_records(user_name)
        except Exception as exc:
            st.error(str(exc))
        else:
            st.success(t("performance.backfill_done").format(count=count))
            st.rerun()

    records = get_production_performance_records()
    if records.empty:
        st.info(t("performance.no_pending"))
        return

    tab_pending, tab_archived = st.tabs([t("performance.pending"), t("performance.archived")])
    with tab_pending:
        pending = records[records["Status"].astype(str).str.casefold().isin(["", "pending"])].copy()
        st.dataframe(_performance_display_frame(pending), use_container_width=True, hide_index=True)
        if pending.empty:
            st.info(t("performance.no_pending"))
        else:
            labels = [_performance_record_label(row) for _, row in pending.iterrows()]
            selected = st.selectbox(t("performance.select_record"), labels, key="performance_pending_select")
            row = pending.iloc[labels.index(selected)].to_dict()
            _performance_editor(row, user_name, "performance_pending_form", allow_archive=True)

    with tab_archived:
        archived = records[records["Status"].astype(str).str.casefold().eq("completed")].copy()
        month_filter = st.text_input("Completed month YYYY-MM", value=pd.Timestamp.now().strftime("%Y-%m"))
        if month_filter:
            completed_dates = pd.to_datetime(archived.get("CompletedAt", ""), errors="coerce")
            archived = archived[completed_dates.dt.strftime("%Y-%m").fillna("").eq(month_filter)]
        st.dataframe(_performance_display_frame(archived), use_container_width=True, hide_index=True)
        if archived.empty:
            st.info(t("performance.no_archived"))
        else:
            labels = [_performance_record_label(row) for _, row in archived.iterrows()]
            selected = st.selectbox(t("performance.select_record"), labels, key="performance_archived_select")
            row = archived.iloc[labels.index(selected)].to_dict()
            reopen_reason = st.text_input(t("performance.reopen_reason"), key="performance_reopen_reason")
            if st.button(t("performance.reopen"), type="secondary"):
                if not reopen_reason.strip():
                    st.error(t("performance.reopen_reason"))
                else:
                    try:
                        update_production_performance_record(
                            user_name,
                            str(row.get("ProductionPlanID", "")),
                            row,
                            reopen_reason=reopen_reason,
                        )
                    except Exception as exc:
                        st.error(str(exc))
                    else:
                        st.success(t("performance.reopen_saved"))
                        st.rerun()


def monthly_report_panel() -> None:
    st.subheader("Monthly Excel Report / 月度整合报告")
    st.caption(
        "Generate a local Excel workbook for production, stock-in, inventory, material consumption, "
        "label consumption, loose goods, and data exceptions. Email sending will be added after the report format is confirmed."
    )

    default_month = default_report_month()
    default_year = int(default_month[:4])
    default_month_num = int(default_month[5:7])
    col_year, col_month = st.columns([1, 1])
    with col_year:
        year = st.number_input("Year / 年份", min_value=2020, max_value=2100, value=default_year, step=1)
    with col_month:
        month = st.selectbox(
            "Month / 月份",
            list(range(1, 13)),
            index=max(default_month_num - 1, 0),
            format_func=lambda value: f"{int(value):02d}",
        )

    report_month = f"{int(year):04d}-{int(month):02d}"
    st.info(f"Report will be generated for / 将生成月份: {report_month}")

    if st.button("Generate Monthly Excel Report / 生成月度 Excel 报告", type="primary"):
        try:
            result = generate_monthly_report(report_month, created_by=current_user().get("username", ""))
        except Exception as exc:
            st.error(f"Failed to generate monthly report: {exc}")
        else:
            st.session_state["monthly_report_result"] = {
                "report_month": result["report_month"],
                "path": str(result["path"]),
                "filename": result["filename"],
                "bytes": result["bytes"],
                "summary_metrics": result["summary_metrics"],
            }
            st.success(f"Monthly report saved: {result['path']}")

    result = st.session_state.get("monthly_report_result")
    if result:
        st.markdown("#### Last generated report / 最近生成的报告")
        st.write(result.get("path", ""))
        summary = pd.DataFrame(result.get("summary_metrics", []))
        if not summary.empty:
            st.dataframe(summary, use_container_width=True, hide_index=True)
        st.download_button(
            "Download Excel Report / 下载 Excel 报告",
            data=result.get("bytes", b""),
            file_name=result.get("filename", f"FactoryMIS_Monthly_Report_{report_month}.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )



def parameter_photo_review_panel() -> None:
    st.subheader("Parameter Adjustment Photos / 参数调整照片")
    st.caption(
        "Cloud uploads are downloaded by the local sync worker. This page is only for local review, notes, and future ML summary data."
    )
    records = get_parameter_photo_records()
    if records.empty:
        st.info("No downloaded parameter adjustment photos yet.")
        return

    status_filter = st.selectbox("Review status", ["Pending", "Confirmed", "All"])
    filtered = records.copy()
    if status_filter != "All" and "ReviewStatus" in filtered.columns:
        filtered = filtered[filtered["ReviewStatus"].astype(str).str.casefold().eq(status_filter.casefold())]
    if filtered.empty:
        st.info("No records match this status.")
        return

    display_columns = [
        "CreatedAt",
        "MachineID",
        "ProductName",
        "MouldNumber",
        "ProblemDescription",
        "UploadedBy",
        "ReviewStatus",
        "DownloadedAt",
    ]
    st.dataframe(filtered[[column for column in display_columns if column in filtered.columns]], use_container_width=True, hide_index=True)

    choices = [
        f"{row.RecordID} | {row.MachineID} | {row.ProductName or row.ProductCode or '-'} | {row.CreatedAt or '-'}"
        for row in filtered.itertuples()
    ]
    selected = st.selectbox("Select record to review", choices)
    row = filtered.iloc[choices.index(selected)].to_dict()

    st.markdown(
        f"**Machine:** {escape(str(row.get('MachineID', '') or '-'))}  \n"
        f"**Product:** {escape(str(row.get('ProductName', '') or row.get('ProductCode', '') or '-'))}  \n"
        f"**Mould:** {escape(str(row.get('MouldNumber', '') or '-'))}  \n"
        f"**Problem:** {escape(str(row.get('ProblemDescription', '') or '-'))}"
    )

    col_before, col_after = st.columns(2)
    for column, title, container in [
        ("BeforePhotoLocalPath", "Before adjustment / 调整前", col_before),
        ("AfterPhotoLocalPath", "After adjustment / 调整后", col_after),
    ]:
        with container:
            st.markdown(f"**{title}**")
            photo_path = Path(str(row.get(column, "") or ""))
            if photo_path.exists():
                st.image(str(photo_path), use_column_width=True)
                st.caption(str(photo_path))
            else:
                st.warning(f"Local file not found: {photo_path}")

    with st.form(f"parameter_photo_review_{row.get('RecordID')}"):
        local_summary = st.text_area(
            "Local summary for future ML / 本地总结（未来机器学习用）",
            value=str(row.get("LocalSummary", "") or ""),
            height=120,
        )
        ml_tags = st.text_input(
            "ML tags / 机器学习标签",
            value=str(row.get("MLTags", "") or ""),
            placeholder="pressure, cooling, colour issue...",
        )
        notes = st.text_area("Local notes / 本地备注", value=str(row.get("Notes", "") or ""), height=80)
        confirmed = st.form_submit_button("Mark reviewed / 标记已确认", type="primary")
    if confirmed:
        update_parameter_photo_review(
            str(row.get("RecordID") or ""),
            current_user()["username"],
            local_summary=local_summary,
            ml_tags=ml_tags,
            notes=notes,
            review_status="Confirmed",
        )
        st.success("Parameter photo record confirmed locally.")
        st.rerun()


def admin_page() -> None:
    st.title("Administrator")
    if not is_admin():
        st.error("Administrator access is required.")
        return

    production = get_production()
    machine_ids = registered_machine_ids(production)
    tab_names = ["Production Schedules", "Production Change Review", "Monthly Reports", "QR Codes", "Parameter Photos", "Data Quality"]
    if can_manage_users():
        tab_names.append("Users")
    tabs = st.tabs(tab_names)

    with tabs[0]:
        if not can_edit_production():
            st.info("Your account does not have permission to modify machine or production schedules.")
        elif production.empty:
            st.warning("No production schedules found.")
        else:
            choices = [
                f"{row.ScheduleID or '(no id)'} | {row.MachineID} | #{row.Sequence or '-'} | {row.ProductName or row.ProductCode or '(empty)'}"
                for row in production.itertuples()
            ]
            selected_choice = st.selectbox("Schedule item to update", choices)
            row_index = choices.index(selected_choice)
            row = production.iloc[row_index].to_dict()
            selected_machine = row.get("MachineID", "")
            current_schedule_id = row.get("ScheduleID", "")

            with st.form("production_form"):
                machine_id = st.selectbox("Machine", machine_ids, index=machine_ids.index(selected_machine) if selected_machine in machine_ids else 0)
                status_options = status_options_for_machine(
                    production,
                    machine_id,
                    ["Running", "Next", "Planned", "Paused", "Finished", "Idle", "Completed"],
                    current_schedule_id,
                )
                show_running_warning_if_needed(production, machine_id, current_schedule_id)
                machine_name = st.text_input("Machine name", value=row.get("MachineName", ""))
                sequence = st.number_input("Sequence", min_value=1, step=1, value=max(int_value(row.get("Sequence")), 1))
                status = st.selectbox(
                    "Status",
                    status_options,
                    index=status_options.index(row.get("Status")) if row.get("Status") in status_options else 0,
                    format_func=production_status_label,
                )
                product_code = st.text_input("Product code / item", value=row.get("ProductCode", ""))
                product_name = st.text_input("Product name", value=row.get("ProductName", ""))
                product_type = st.text_input("Product type", value=row.get("ProductType", ""))
                size = st.text_input("Size", value=row.get("Size", ""))
                planned = st.number_input("Planned quantity", min_value=0, step=1, value=int_value(row.get("PlannedQty")))
                completed = st.number_input("Completed quantity", min_value=0, step=1, value=int_value(row.get("CompletedQty")))
                qc_notes = st.text_area(
                    t("production.qc_details"),
                    value=str(row.get("QCNotes", "") or ""),
                    height=80,
                    placeholder=t("production.qc_details_placeholder"),
                    key="admin_qc_notes",
                )
                mould_number = st.text_input("Mould number", value=row.get("MouldNumber", ""))
                material = st.text_input("Material", value=row.get("Material", ""))
                material_location = st.text_input("Material location", value=row.get("MaterialLocation", ""))
                colour = st.text_input("Colour / masterbatch", value=row.get("ColourMasterbatch", ""))
                material_weight = st.text_input("Main material weight (g/pc)", value=row.get("MaterialWeightG", ""))
                second_material_weight = st.text_input("Second material weight (g/pc)", value=row.get("SecondMaterialWeightG", ""))
                masterbatch_weight = st.text_input("Masterbatch weight (g/pc)", value=row.get("MasterbatchWeightG", ""))
                unit_weight = st.text_input("Unit total weight (g/pc)", value=row.get("UnitWeightG", ""))
                label = st.text_input("Label", value=row.get("Label", ""))
                packaging_type = st.text_input("Packaging type", value=row.get("PackagingType", ""))
                packaging_option = st.text_input("Packaging option", value=row.get("PackagingOption", ""))
                packaging_unit = st.text_input("Packaging unit", value=row.get("PackagingUnit", ""))
                carton_stack = st.text_input("Carton qty / unit / stack", value=row.get("CartonUnitStackQty", ""))
                pallet_type = st.text_input("Pallet type", value=row.get("PalletType", ""))
                pallet_qty = st.text_input("Pallet qty", value=row.get("PalletQty", ""))
                notes = st.text_area("Production notes", value=row.get("Notes", ""))
                submitted = st.form_submit_button("Save Production Schedule")

            if submitted:
                update_production_record(
                    current_user()["username"],
                    machine_id,
                    {
                        "MachineName": machine_name,
                        "Sequence": str(sequence),
                        "Status": status,
                        "ProductType": product_type,
                        "Size": size,
                        "ProductCode": product_code,
                        "ProductName": product_name,
                        "PlannedQty": str(planned),
                        "CompletedQty": str(completed),
                        "MouldNumber": mould_number,
                        "QCRequired": "YES" if qc_required_for_save(qc_notes) else "NO",
                        "QCNotes": qc_notes_for_save(qc_notes),
                        "Material": material,
                        "MaterialLocation": material_location,
                        "ColourMasterbatch": colour,
                        "MaterialWeightG": material_weight,
                        "SecondMaterialWeightG": second_material_weight,
                        "MasterbatchWeightG": masterbatch_weight,
                        "UnitWeightG": unit_weight,
                        "Label": label,
                        "PackagingUnit": packaging_unit,
                        "PackagingType": packaging_type,
                        "PackagingOption": packaging_option,
                        "CartonUnitStackQty": carton_stack,
                        "PalletQty": pallet_qty,
                        "AdditionalPackaging": row.get("AdditionalPackaging", ""),
                        "PalletBag": row.get("PalletBag", ""),
                        "PalletType": pallet_type,
                        "WrapPallet": row.get("WrapPallet", ""),
                        "FoodApplication": row.get("FoodApplication", ""),
                        "CornerProtector": row.get("CornerProtector", ""),
                        "InventoryLocationID": row.get("InventoryLocationID", ""),
                        "AdditionalInstructions": row.get("AdditionalInstructions", ""),
                        "Notes": notes,
                    },
                    schedule_id=row.get("ScheduleID", ""),
                )
                st.success("Production schedule saved.")
                st.rerun()

            with st.expander("Delete selected production item", expanded=False):
                schedule_id = str(row.get("ScheduleID", ""))
                st.warning("Deletion is permanent in ProductionSchedule.xlsx and will be recorded in ChangeLog.xlsx.")
                confirm_delete = st.checkbox(
                    f"Confirm deletion of {schedule_id}",
                    key=f"admin_confirm_delete_{schedule_id}",
                )
                if st.button(
                    "Delete Selected Item",
                    key=f"admin_delete_schedule_{schedule_id}",
                    disabled=not confirm_delete,
                ):
                    delete_production_record(current_user()["username"], schedule_id)
                    st.success("Production schedule deleted.")
                    st.rerun()

    with tabs[1]:
        production_change_review_panel()

    with tabs[2]:
        monthly_report_panel()

    with tabs[3]:
        qr_generator_panel(machine_ids)

    with tabs[4]:
        parameter_photo_review_panel()

    with tabs[5]:
        data_quality_panel()

    if can_manage_users():
        with tabs[6]:
            user_management_panel()


def now_file_stamp() -> str:
    return pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")




def is_label_value(value: object) -> bool:
    text = str(value or "").strip()
    return text.upper() not in {"", "NO", "N", "FALSE", "0", "NONE", "N/A", "NA", "-", "NO LABEL"}


LABEL_INVENTORY_COLUMNS = [
    "LabelKey",
    "Brand",
    "Size",
    "ProductType",
    "LabelDisplayName",
    "LabelName",
    "CurrentStock",
    "Location",
    "Notes",
    "UpdatedAt",
]


def label_inventory_path() -> Path:
    return Path(__file__).resolve().parent / "data" / "LabelInventory.xlsx"


def current_label_quarter() -> str:
    now = pd.Timestamp.now()
    quarter = ((int(now.month) - 1) // 3) + 1
    return f"{int(now.year)}-Q{quarter}"


def quarter_sort_key(value: object) -> tuple[int, int]:
    match = re.match(r"^(\d{4})-Q([1-4])$", str(value or "").strip())
    if not match:
        return (0, 0)
    return (int(match.group(1)), int(match.group(2)))


def timestamp_to_label_quarter(value: object) -> str:
    timestamp = pd.to_datetime(value, errors="coerce")
    if pd.isna(timestamp):
        return ""
    quarter = ((int(timestamp.month) - 1) // 3) + 1
    return f"{int(timestamp.year)}-Q{quarter}"


def available_label_quarters() -> list[str]:
    quarters = {current_label_quarter()}
    history = get_stock_history()
    if not history.empty and "Timestamp" in history.columns:
        for value in history["Timestamp"].dropna().astype(str):
            quarter = timestamp_to_label_quarter(value)
            if quarter:
                quarters.add(quarter)
    return sorted(quarters, key=quarter_sort_key, reverse=True)


def label_brand_from_catalog_row(row: pd.Series) -> str:
    text = f"{row.get('Item', '')} {row.get('HasLabel', '')}".upper()
    compact = re.sub(r"[^A-Z0-9]", "", text)
    if "CROM" in compact:
        return "Crommelin"
    if "TRIMLAGNEWS" in compact or ("AGNEWS" in compact and "TIMBM" not in compact and "TIMBERMATE" not in compact):
        return "AGNEWS"
    if "AMGROW" in compact:
        return "Amgrow"
    if "TIMBM" in compact or "TIMBERMATE" in compact:
        return "Timbermate"
    return "Other"


def label_size_sort_value(size: object) -> float:
    text = str(size or "").upper().replace(" ", "")
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    if not match:
        return 999999.0
    value = float(match.group(1))
    if "ML" in text:
        return value / 1000.0
    if "L" in text:
        return value
    if "KG" in text:
        return 10000.0 + value
    return 50000.0 + value


def normalize_label_token_text(value: object) -> str:
    text = str(value or "").upper()
    text = text.replace("JAY'S", "JAYS")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def label_display_name_from_product_code(row: pd.Series, brand: str) -> str:
    raw = str(row.get("Item", "") or "").upper()
    raw = raw.replace("JAY'S", "JAYS")
    size = str(row.get("Size", "") or "").upper().replace(" ", "")
    if brand == "AGNEWS" and "AGNEWS" in raw.replace(" ", ""):
        return "AGNEWS"

    text = raw
    replacements = [
        "TRIMLAGNEWS",
        "TIMBERMATE",
        "TIMBM",
        "CROMMELIN",
        "CROM",
        "AMGROW",
        "IM LABEL",
        "IML",
        "IMC",
        "LABEL",
        "BUCKET",
        "BUC",
        "PAIL",
        "TUB",
        "LID",
    ]
    if size:
        replacements.extend([size, size.replace("L", "LTR"), size.replace("L", "LT")])
        if size.endswith("ML"):
            replacements.append(size[:-2] + " ML")
        if size.endswith("L"):
            replacements.append(size[:-1] + " L")
    for token in sorted(set(replacements), key=len, reverse=True):
        if token:
            text = text.replace(token, " ")
    text = normalize_label_token_text(text)
    text = re.sub(r"\b[1-9]\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or str(row.get("HasLabel", "") or row.get("Item", "") or "").strip()


def label_key(*parts: object) -> str:
    return "|".join(normalize_label_token_text(part).casefold() for part in parts)


def _history_product_key(row: pd.Series) -> str:
    product_code = str(row.get("ProductCode", "") or "").strip()
    product_name = str(row.get("ProductName", "") or "").strip()
    return (product_code or product_name).casefold()


def read_label_inventory() -> pd.DataFrame:
    path = label_inventory_path()
    if not path.exists():
        return pd.DataFrame(columns=LABEL_INVENTORY_COLUMNS)
    inventory = pd.read_excel(path, dtype=str).fillna("")
    for column in LABEL_INVENTORY_COLUMNS:
        if column not in inventory.columns:
            inventory[column] = ""
    return inventory[LABEL_INVENTORY_COLUMNS]


def build_label_catalog_groups() -> tuple[pd.DataFrame, dict[str, str]]:
    catalog = get_product_catalog().copy()
    if "HasLabel" not in catalog.columns:
        return pd.DataFrame(), {}
    labelled = catalog[catalog["HasLabel"].apply(is_label_value)].copy()
    if labelled.empty:
        return pd.DataFrame(), {}

    groups: dict[str, dict[str, object]] = {}
    product_to_group: dict[str, str] = {}
    brand_order = {"Crommelin": 1, "Timbermate": 2, "AGNEWS": 3, "Amgrow": 4, "Other": 99}
    for _, row in labelled.iterrows():
        brand = label_brand_from_catalog_row(row)
        size = str(row.get("Size", "") or "").strip()
        product_type = str(row.get("ProductType", "") or "").strip()
        display_name = label_display_name_from_product_code(row, brand)
        label_name = str(row.get("HasLabel", "") or "").strip()
        key = label_key(brand, size, product_type, display_name, label_name)
        item = str(row.get("Item", "") or "").strip()
        if key not in groups:
            groups[key] = {
                "LabelKey": key,
                "Brand": brand,
                "Size": size,
                "ProductType": product_type,
                "LabelDisplayName": display_name,
                "LabelName": label_name,
                "ProductCodes": set(),
                "_BrandOrder": brand_order.get(brand, 99),
                "_SizeSort": label_size_sort_value(size),
            }
        if item:
            groups[key]["ProductCodes"].add(item)
            product_to_group[item.casefold()] = key
    rows = []
    for group in groups.values():
        row = group.copy()
        row["ProductCodes"] = ", ".join(sorted(row["ProductCodes"]))
        rows.append(row)
    result = pd.DataFrame(rows)
    result = result.sort_values(["_BrandOrder", "Brand", "_SizeSort", "Size", "ProductType", "LabelDisplayName"]).reset_index(drop=True)
    return result, product_to_group


def build_label_consumption_log(quarter: str | None = None) -> pd.DataFrame:
    selected_quarter = quarter or current_label_quarter()
    groups, product_to_group = build_label_catalog_groups()
    if groups.empty:
        return pd.DataFrame()

    consumed = {str(key): 0.0 for key in groups["LabelKey"]}
    history = get_stock_history().copy()
    if not history.empty:
        for column in ["ProductCode", "ProductName", "Quantity", "Timestamp"]:
            if column not in history.columns:
                history[column] = ""
        history["_QuantityNum"] = pd.to_numeric(history["Quantity"], errors="coerce").fillna(0)
        history["_Quarter"] = history["Timestamp"].apply(timestamp_to_label_quarter)
        history = history[history["_Quarter"].astype(str) == str(selected_quarter)]
        for _, hist_row in history.iterrows():
            product_key = _history_product_key(hist_row)
            group_key = product_to_group.get(product_key)
            if not group_key:
                continue
            qty = float(hist_row.get("_QuantityNum") or 0)
            if qty > 0:
                consumed[group_key] = consumed.get(group_key, 0.0) + qty

    inventory = read_label_inventory()
    stock_lookup = dict(zip(inventory["LabelKey"].astype(str), inventory["CurrentStock"].astype(str))) if not inventory.empty else {}
    location_lookup = dict(zip(inventory["LabelKey"].astype(str), inventory["Location"].astype(str))) if not inventory.empty else {}
    notes_lookup = dict(zip(inventory["LabelKey"].astype(str), inventory["Notes"].astype(str))) if not inventory.empty else {}

    rows: list[dict[str, object]] = []
    for _, group in groups.iterrows():
        key = str(group.get("LabelKey", ""))
        qty = consumed.get(key, 0.0)
        qty_display: int | float = int(qty) if float(qty).is_integer() else qty
        rows.append(
            {
                "Quarter": selected_quarter,
                "LabelKey": key,
                "Brand": group.get("Brand", ""),
                "Size": group.get("Size", ""),
                "ProductType": group.get("ProductType", ""),
                "LabelDisplayName": group.get("LabelDisplayName", ""),
                "LabelName": group.get("LabelName", ""),
                "LabelConsumedQty": qty_display,
                "CurrentStock": stock_lookup.get(key, ""),
                "Location": location_lookup.get(key, ""),
                "Notes": notes_lookup.get(key, ""),
                "ProductCodes": group.get("ProductCodes", ""),
            }
        )
    return pd.DataFrame(rows)


def label_display_dataframe(label_df: pd.DataFrame) -> pd.DataFrame:
    display_columns = [
        "Brand",
        "Size",
        "ProductType",
        "LabelDisplayName",
        "LabelName",
        "LabelConsumedQty",
        "CurrentStock",
    ]
    display = label_df.copy()
    for column in display_columns:
        if column not in display.columns:
            display[column] = ""
    return display[display_columns].rename(
        columns={
            "Brand": "\u54c1\u724c",
            "Size": "\u5927\u5c0f",
            "ProductType": "\u4ea7\u54c1\u7c7b\u578b",
            "LabelDisplayName": "\u6807\u7b7e\u540d\u79f0",
            "LabelName": "LabelName",
            "LabelConsumedQty": "\u6807\u7b7e\u7d2f\u8ba1\u6d88\u8017",
            "CurrentStock": "\u73b0\u6709\u5e93\u5b58",
        }
    )


def label_consumption_workbook(label_df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font, PatternFill

    output = BytesIO()
    export_df = label_display_dataframe(label_df)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Label Consumption")
        for worksheet in writer.sheets.values():
            worksheet.freeze_panes = "A2"
            header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            widths = {"A": 16, "B": 10, "C": 12, "D": 26, "E": 44, "F": 16, "G": 14}
            for column_letter, width in widths.items():
                worksheet.column_dimensions[column_letter].width = width
    return output.getvalue()

def machine_archive_workbook(archive_df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font, PatternFill

    output = BytesIO()
    export_df = archive_df.copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Machine Archive")
        worksheet = writer.sheets["Machine Archive"]
        worksheet.freeze_panes = "A2"
        widths = {
            "A": 20,
            "B": 22,
            "C": 12,
            "D": 16,
            "E": 14,
            "H": 18,
            "I": 34,
            "J": 12,
            "K": 14,
            "L": 14,
            "N": 16,
            "O": 18,
            "Q": 24,
            "Y": 24,
            "Z": 40,
            "AA": 40,
        }
        for column_letter, width in widths.items():
            worksheet.column_dimensions[column_letter].width = width
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        complete_fill = PatternFill(fill_type="solid", fgColor="D9F99D")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in worksheet.iter_rows(min_row=2):
            if len(row) > 6:
                row[6].fill = complete_fill
    return output.getvalue()


def history_page() -> None:
    st.title("History")
    tab_changes, tab_stock, tab_label, tab_loose, tab_machine_archive = st.tabs(
        [
            "Change Log",
            t("history.stock_in"),
            "Label Consumption / \u6807\u7b7e\u6d88\u8017",
            t("history.loose_goods"),
            "Machine Archive / \u673a\u5668\u5f52\u6863",
        ]
    )

    with tab_changes:
        changes = get_change_log()
        keyword = st.text_input("Search change history", key="change_search")
        if keyword:
            mask = pd.Series(False, index=changes.index)
            for column in changes.columns:
                mask = mask | changes[column].astype(str).str.contains(keyword, case=False, na=False)
            changes = changes[mask]
        if changes.empty:
            st.info("No change history records match the current search.")
        else:
            st.dataframe(changes.sort_index(ascending=False), use_container_width=True, hide_index=True)

    with tab_stock:
        stock_history = get_stock_history()
        keyword = st.text_input(t("history.search"), key="stock_history_search", placeholder="Product, machine, operator, request id")
        if st.session_state.get("stock_history_search_last") != keyword:
            st.session_state["stock_history_search_last"] = keyword
            st.session_state["stock_history_offset"] = 0
        if keyword:
            mask = pd.Series(False, index=stock_history.index)
            for column in stock_history.columns:
                mask = mask | stock_history[column].astype(str).str.contains(keyword, case=False, na=False)
            stock_history = stock_history[mask]
        show_last = st.slider(t("history.show_last"), 10, 200, 10, step=10, key="stock_history_limit")
        max_offset = max((len(stock_history) - 1) // show_last, 0) if show_last else 0
        if int(st.session_state.get("stock_history_offset", 0) or 0) > max_offset:
            st.session_state["stock_history_offset"] = max_offset
        offset = st.slider(t("history.page_offset"), 0, max_offset, 0, key="stock_history_offset") if max_offset else 0
        offset = min(offset, max_offset)
        stock_history = stock_history.sort_index(ascending=False)
        start = offset * show_last
        end = start + show_last
        display_columns = [
            "Timestamp", "User", "ProductCode", "ProductName", "Quantity",
            "OldStock", "NewStock", "MachineID", "ScheduleID", "MouldNumber",
            "QuantityMode", "RequestType", "ClientRequestID", "ProcessedBy",
            "ProcessedTime", "Remarks",
        ]
        for column in display_columns:
            if column not in stock_history.columns:
                stock_history[column] = ""
        if stock_history.empty:
            st.info("No stock-in history records match the current search.")
        else:
            st.caption(f"Showing records {start + 1}-{min(end, len(stock_history))} of {len(stock_history)}")
            st.dataframe(stock_history.iloc[start:end][display_columns], use_container_width=True, hide_index=True)


    with tab_label:
        quarters = available_label_quarters()
        default_quarter = current_label_quarter()
        default_index = quarters.index(default_quarter) if default_quarter in quarters else 0
        selected_quarter = st.selectbox(
            "Quarter / \u5b63\u5ea6",
            quarters,
            index=default_index,
            key="label_consumption_quarter",
            help="Default is the current quarter. Choose an older quarter to query history.",
        )
        label_log = build_label_consumption_log(selected_quarter)
        if label_log.empty:
            st.info("No labelled products found in ProductCatalog.xlsx yet.")
        else:
            brand_options = list(label_log["Brand"].dropna().astype(str).drop_duplicates())
            if not brand_options:
                st.info("No label brands found for the selected quarter.")
            else:
                brand_filter = st.selectbox("Brand / \u54c1\u724c", brand_options, key="label_consumption_brand")
                brand_log = label_log[label_log["Brand"].astype(str) == brand_filter].copy()
                size_options = sorted(
                    brand_log["Size"].dropna().astype(str).drop_duplicates(),
                    key=label_size_sort_value,
                )
                if not size_options:
                    st.info("No label sizes found for the selected brand.")
                else:
                    size_filter = st.selectbox("Size / \u5927\u5c0f", size_options, key="label_consumption_size")
                    filtered = brand_log[brand_log["Size"].astype(str) == size_filter].copy()
                    selected_consumed = pd.to_numeric(filtered["LabelConsumedQty"], errors="coerce").fillna(0).sum()
                    selected_consumed_display = int(selected_consumed) if float(selected_consumed).is_integer() else selected_consumed
                    st.metric(
                        f"{selected_quarter} {brand_filter} {size_filter} label consumption / \u5f53\u524d\u7b5b\u9009\u6807\u7b7e\u6d88\u8017",
                        selected_consumed_display,
                    )

                    st.caption(
                        "Consumption follows the selected brand and size only. Current stock is reserved for label purchase/stock updates. / "
                        "\u6d88\u8017\u53ea\u6309\u5df2\u9009\u54c1\u724c\u548c\u5c3a\u5bf8\u8ba1\u7b97\uff1b\u73b0\u6709\u5e93\u5b58\u7528\u4e8e\u540e\u7eed\u6807\u7b7e\u91c7\u8d2d\u5165\u5e93\u8bb0\u5f55\u3002"
                    )
                    st.dataframe(label_display_dataframe(filtered), use_container_width=True, hide_index=True)
                    st.download_button(
                        "Download current view Excel / \u4e0b\u8f7d\u5f53\u524d\u5217\u8868 Excel",
                        data=label_consumption_workbook(filtered),
                        file_name=f"label_consumption_{selected_quarter}_{brand_filter}_{size_filter}_{now_file_stamp()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )


    with tab_loose:
        loose_history = get_loose_goods()
        keyword = st.text_input(t("history.search"), key="loose_history_search", placeholder="Loose ID, product, machine, operator, request id")
        if st.session_state.get("loose_history_search_last") != keyword:
            st.session_state["loose_history_search_last"] = keyword
            st.session_state["loose_history_offset"] = 0
        if keyword:
            mask = pd.Series(False, index=loose_history.index)
            for column in loose_history.columns:
                mask = mask | loose_history[column].astype(str).str.contains(keyword, case=False, na=False)
            loose_history = loose_history[mask]
        show_last = st.slider(t("history.show_last"), 10, 200, 10, step=10, key="loose_history_limit")
        max_offset = max((len(loose_history) - 1) // show_last, 0) if show_last else 0
        if int(st.session_state.get("loose_history_offset", 0) or 0) > max_offset:
            st.session_state["loose_history_offset"] = max_offset
        offset = st.slider(t("history.page_offset"), 0, max_offset, 0, key="loose_history_offset") if max_offset else 0
        offset = min(offset, max_offset)
        loose_history = loose_history.sort_index(ascending=False)
        start = offset * show_last
        end = start + show_last
        if loose_history.empty:
            st.info("No loose goods records match the current search.")
        else:
            st.caption(f"Showing records {start + 1}-{min(end, len(loose_history))} of {len(loose_history)}")
            st.dataframe(loose_history.iloc[start:end], use_container_width=True, hide_index=True)

    with tab_machine_archive:
        archive_df = get_machine_archive()
        keyword = st.text_input(
            "Search machine archive / 搜索机器归档",
            key="machine_archive_search",
            placeholder="Machine, product, schedule, mould, material, colour",
        )
        if keyword:
            mask = pd.Series(False, index=archive_df.index)
            for column in archive_df.columns:
                mask = mask | archive_df[column].astype(str).str.contains(keyword, case=False, na=False)
            archive_df = archive_df[mask]
        if archive_df.empty:
            st.info("No archived machine production records yet.")
        else:
            display_columns = [
                "ArchivedAt", "ArchiveReason", "MachineID", "MachineName", "ScheduleID",
                "Sequence", "Status", "ProductCode", "ProductName", "PlannedQty",
                "CompletedQty", "FinalRemainingQty", "FinalProgressPercent",
                "MouldNumber", "Material", "MaterialLocation", "ColourMasterbatch",
                "PalletQty", "PackagingUnit", "PackagingType", "CartonUnitStackQty",
                "PalletBag", "PalletType", "WrapPallet", "CornerProtector",
                "InventoryLocationID", "AdditionalInstructions", "Notes", "LastUpdated",
            ]
            for column in display_columns:
                if column not in archive_df.columns:
                    archive_df[column] = ""
            archive_df = archive_df.sort_values("ArchivedAt", ascending=False)
            st.dataframe(archive_df[display_columns], use_container_width=True, hide_index=True)
            st.download_button(
                "Download Machine Archive Excel / 下载机器归档 Excel",
                data=machine_archive_workbook(archive_df[display_columns]),
                file_name=f"machine_archive_{now_file_stamp()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


def main() -> None:
    inject_css()
    inject_shared_theme()
    requested_page = query_value("page", "machine")
    machine_id = query_value("machine_id", "")
    public_mode = query_flag("public")

    if public_mode and requested_page == "machine":
        public_top_bar()
        if machine_id:
            machine_page(machine_id, public_view=True)
        else:
            public_machine_overview()
        public_bottom_nav("machine")
        return

    login_panel()

    page = nav(requested_page)

    if page == "machine":
        machine_page(machine_id)
    elif page == "production_table":
        production_table_page()
    elif page == "stock_in":
        stock_in_page()
    elif page == "loose_goods":
        loose_goods_page()
    elif page == "moulds":
        mould_page()
    elif page == "product_mould_links":
        product_mould_links_page()
    elif page == "history":
        history_page()
    elif page == "admin":
        admin_page()


if __name__ == "__main__":
    main()
